"""Apples & Ginos 2026-27 NHL Skater Projections - Nate -- same layout as
apples_ginos_blake.py (see that module's docstring for the header-collision/broken-Pos/
missing-category details); only the filename and sheet name differ. Filename genuinely
starts with an underscore in Sheets/, not a typo introduced here.
"""

from pathlib import Path

import openpyxl

FILENAME = "_Apples & Ginos 2026-27 NHL Skater Projections - Nate.xlsx"
SHEET = "Nates Projections"
HEADER_ROW = 7
FIRST_DATA_ROW = 8


def _to_int(value):
    return int(round(float(value))) if value not in (None, "") else None


def _to_float(value):
    return float(value) if value not in (None, "") else None


def _header_map(ws, row):
    out = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=c).value
        if isinstance(v, str) and v.strip() and v.strip() not in out:
            out[v.strip()] = c
    return out


def rows(sheets_dir: Path):
    wb = openpyxl.load_workbook(sheets_dir / FILENAME, data_only=True)
    ws = wb[SHEET]
    hmap = _header_map(ws, HEADER_ROW)

    def get(row, header):
        return row[hmap[header] - 1]

    for r in ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
        name = get(r, "Name")
        if not name:
            continue
        yield {
            "raw_name": name,
            "team_raw": get(r, "Team"),
            "is_goalie": False,
            "stats": {
                "GamesPlayed": _to_int(get(r, "GP")),
                "Goals": _to_int(get(r, "G")),
                "Assists": _to_int(get(r, "A")),
                "Points": _to_int(get(r, "PTS")),
                "PowerPlayPoints": _to_int(get(r, "PPP")),
                "Shots": _to_int(get(r, "SOG")),
                "Hits": _to_int(get(r, "HIT")),
                "Blocks": _to_int(get(r, "BLK")),
                "PenaltyMinutes": _to_int(get(r, "PIM")),
                "AverageTOIMinutes": _to_float(get(r, "ATOI")),
            },
        }
