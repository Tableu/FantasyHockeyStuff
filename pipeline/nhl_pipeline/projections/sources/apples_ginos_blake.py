"""Apples & Ginos 2026-27 NHL Skater Projections - Blake -- skaters only, one row per player.
The sheet is a fantasy-points calculator, not a flat stat table: row 7 is the real header,
and its raw season-total projections (columns A:O) are followed by a second, differently
labeled block (columns P onward) holding per-category FANTASY POINT contributions (goals
weighted by the sheet's own scoring, etc.), not raw stats -- that second block reuses the
same header text (G, A, PTS, PPP, SOG, HIT, BLK, PIM, S%, ATOI), so header_map keeps only the
first occurrence of each name, which lands on the real raw-stat block. No Y! Pos values are
usable (every row reads '#REF!', a broken formula in the source file) and no FOW/FOL/+/-/
PPG/PPA columns exist at all, so those stay unset for this source.
"""

from pathlib import Path

import openpyxl

FILENAME = "Apples & Ginos 2026-27 NHL Skater Projections - Blake.xlsx"
SHEET = "Blakes Projections"
HEADER_ROW = 7
FIRST_DATA_ROW = 8


def _to_int(value):
    return int(round(float(value))) if value not in (None, "") else None


def _to_float(value):
    return float(value) if value not in (None, "") else None


def _header_map(ws, row):
    out = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=c).value
        if isinstance(v, str) and v.strip() and v.strip() not in out:
            out[v.strip()] = c
    return out


def rows(sheets_dir: Path):
    wb = openpyxl.load_workbook(sheets_dir / FILENAME, data_only=True)
    ws = wb[SHEET]
    hmap = _header_map(ws, HEADER_ROW)

    def get(row, header):
        return row[hmap[header] - 1]

    for r in ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
        name = get(r, "Name")
        if not name:
            continue
        yield {
            "raw_name": name,
            "team_raw": get(r, "Team"),
            "is_goalie": False,
            "stats": {
                "GamesPlayed": _to_int(get(r, "GP")),
                "Goals": _to_int(get(r, "G")),
                "Assists": _to_int(get(r, "A")),
                "Points": _to_int(get(r, "PTS")),
                "PowerPlayPoints": _to_int(get(r, "PPP")),
                "Shots": _to_int(get(r, "SOG")),
                "Hits": _to_int(get(r, "HIT")),
                "Blocks": _to_int(get(r, "BLK")),
                "PenaltyMinutes": _to_int(get(r, "PIM")),
                "AverageTOIMinutes": _to_float(get(r, "ATOI")),
            },
        }
