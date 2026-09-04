#!/usr/bin/env python
"""Builds the 'Aggregate Fantasy Hockey Projections 2026-27' Google Sheet (RESULT) from the
provided projection sources, using the live 'Aggregate Sheet Template (clean)' Google Sheet
(SOURCE) as the structural template. Both are read/written entirely through the Sheets API
(via gsheets_io.py, an adapter exposing just enough of openpyxl's Workbook/Worksheet surface
for the phase functions below to run unmodified) -- no local .xlsx file is involved.

SOURCE holds structure only -- every sheet's formulas, named ranges, and data validation, but
no player/projection/schedule data, and no raw per-source tabs at all (the original template,
still at 1qWbfN9TwDFWLR3bcCFKyIrqV8H7chSkHUgyJ33yjquI, is kept as a fallback but is no longer
what this script bootstraps from). Sheets a human maintains directly and this script never
writes to at all -- NameFix's alias table, MISC3/Fleaflicker's paste-and-check columns, Cheat
Sheet's ARRAY_CONSTRAIN spill, Rankings!B:I, Settings' scoring/roster config -- are exactly as
blank/ready-to-use in SOURCE as they'd be after a human cleared last season's leftovers by
hand; nothing about them changes here. (Import 1/2/3, the old "paste your own projections
here" sheets, are retired -- see DROP_SHEETS -- now that a new source only needs to be
imported into the database to be picked up automatically.)

Reworked (regenerated from scratch every run, not preserved from SOURCE, and self-sufficient
even for a from-scratch bootstrap that never saw these sheets before -- see
ensure_raw_source_sheet/ensure_source_named_ranges): the raw per-source tabs, one per row
Projections.Sources has for the current season (not a hardcoded list -- see _active_sources),
Positions, AllProjections_S/G, Export, ADPYahoo/ADPFantrax/ADPother, Rankings!A, CVals/Vorp,
FanPts/Vorp, CleanCat/CleanPts, Available - Cats/Pts, and Player Values - Cats/Pts' player
rows. Every reworked formula below is plain pre-2007-function arithmetic (SUMPRODUCT/INDEX/
MATCH instead of FILTER/SORT/ARRAY_CONSTRAIN) -- not because native Sheets needs it (it
doesn't), but because it was originally written to also run correctly if re-exported to Excel;
kept as-is since it's already correct and there's no reason to churn it. SORTBY is used
instead of SORT specifically to avoid SORT's sort_order argument meaning opposite things in
Excel (1/-1) vs Sheets (TRUE/FALSE); descending sorts are done by sorting ascending on a
negated key instead.

Player Values - Cats/Pts is a hybrid: the *pattern* each player's row of formulas follows is
authored directly in the sheet (two reference rows -- see rebuild_player_values), not
hardcoded here, but this script still writes the row-shifted result into every real player
row every run (Google Sheets has no live per-row conditional-formula-variant primitive that
would let the sheet do this entirely on its own). Editing what a Player Values formula does
is a template edit now, not a code change; editing how many player rows exist, or which
reference row a given player uses, is still this module's job.

Player-name matching across sources is exact-string (per user: already reconciled).

Usage:
    python build_aggregate_workbook.py
"""

import logging
import re
import sys
from datetime import timedelta
from pathlib import Path

# Makes "python build_aggregate_workbook.py" work regardless of the caller's cwd, matching
# this module's own documented usage -- nhl_pipeline lives one directory up, which is only on
# sys.path automatically when this script is run from the pipeline root itself.
_PIPELINE_ROOT = str(Path(__file__).resolve().parent.parent)
if _PIPELINE_ROOT not in sys.path:
    sys.path.insert(0, _PIPELINE_ROOT)

from openpyxl.utils import column_index_from_string as ci, get_column_letter as cl

import gsheets_io
from gsheets_io import set_defined_name
from nhl_pipeline import db as nhl_db

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("build_aggregate_workbook")

# Aggregate Sheet Template (clean) -- a data-free duplicate of the original template (still at
# 1qWbfN9TwDFWLR3bcCFKyIrqV8H7chSkHUgyJ33yjquI, kept as a fallback). Every sheet's structure,
# named ranges, data validation, and formulas (including the ones this module never touches --
# NameFix's alias dropdown, Rankings!B:I, Cheat Sheet's ARRAY_CONSTRAIN spill, etc.) are
# unchanged; only literal player/projection/schedule DATA was cleared out, so a from-scratch
# bootstrap no longer starts from ~600-1000 rows of leftover sample names in tabs like MISC3/
# Fleaflicker/Import 1-3 that this script was never going to overwrite anyway. Positions' own
# handful of genuinely-dead legacy columns (see the removed POSITIONS_DEAD_COLS) were
# deliberately left in place rather than deleted -- deleting them would shift column T, which
# this module's own POSITIONS_MANUAL_EXTRA_COL now writes to every run, and re-deriving that
# shift correctly is exactly the class of bug _repatch_shifted_positions_refs used to exist to
# paper over. Leaving a few permanently-blank columns costs nothing; getting that shift wrong
# would corrupt fill_positions' own output.
SOURCE_SPREADSHEET_ID = "1B-Vp2edky2VE2eV1QmK37D0805lK2RRBRdNUkW2ip_0"
RESULT_SPREADSHEET_ID = "11zECvxbuKwlYBncjE75VncFLUPkp90fv6LsjoAp-97E"  # Aggregate Fantasy Hockey Projections 2026-27
CREDENTIALS_PATH = Path(__file__).parent / "googleSheetsCredentials.json"

TEAMS = 12  # Settings!C3, used by CleanCat's draft-pick-label formula
SEASON_NHL_ID = 20262027

# "Which sources actually exist" is read from Projections.Sources every run, not hardcoded --
# see _active_sources below, which populates the ACTIVE_SOURCES/DB_EXPORTED_SOURCES/
# ACTIVE_SOURCE_SHEETS globals declared here as soon as load_master_data has a cursor. A source
# that stops being imported drops out on the next run with no code change; one newly imported
# (e.g. a source added mid-season) is picked up the same way -- except anything named in
# DB_SOURCE_EXCLUDE, which the database has real data for but that isn't ready to blend in yet.
# Everything downstream that needs the active source list (rebuild_settings_sources,
# rebuild_source_check, rebuild_source_comparison, rebuild_all_projections,
# rebuild_misc_source_weights) already reads these three names generically -- none of them
# needed to change for this to become database-driven instead of a fixed list.
# Dom (The Athletic)'s export -- imported, not ready to blend in yet (per user instruction).
# Both names covered defensively: SourceID 4's SourceName changed from "Fantrax" to "Dom"
# (same row, same Description) mid-development, so it isn't safe to assume which one is
# current by the time this runs again.
DB_SOURCE_EXCLUDE = {"Fantrax", "Dom"}
ACTIVE_SOURCES = []
DB_EXPORTED_SOURCES = set()
ACTIVE_SOURCE_SHEETS = []

_NICK_RE = re.compile(r"[^A-Za-z0-9]")


def _source_nick(source_name):
    """A short, named-range-safe identifier derived from the source's own database name
    (Sheets named ranges can't contain spaces or punctuation) -- e.g. "Lineup Experts" ->
    "LineupExperts", "Apples & Ginos - Blake" -> "ApplesGinosBlake". Deterministic, so a
    newly-imported source never needs a nick manually assigned to it."""
    return _NICK_RE.sub("", source_name)


def _active_sources(cursor) -> list:
    """[(source_name, source_name, nick), ...] for every source Projections.Sources has real
    data for this season, alphabetical, minus DB_SOURCE_EXCLUDE. source_name doubles as both
    the Settings!S label and the raw-source sheet's own title (see ensure_raw_source_sheet) --
    one name to keep straight per source, not three hand-kept in sync."""
    cursor.execute(
        """
        SELECT DISTINCT src.SourceName
        FROM Projections.Sources src
        JOIN Reference.Seasons se ON se.SeasonID = src.SeasonID
        WHERE se.NHLSeasonID = ?
        ORDER BY src.SourceName
        """,
        SEASON_NHL_ID,
    )
    names = [row.SourceName for row in cursor.fetchall() if row.SourceName not in DB_SOURCE_EXCLUDE]
    return [(name, name, _source_nick(name)) for name in names]


# Import 1/2/3 -- a human pastes their own projections directly into one of these (following
# its header row -- the same unified RAW_SOURCE_HEADERS every source's sheet uses, so a pasted
# stat lands under the same category codes CatsAll/CatWeights already use elsewhere, e.g. "P"
# not "PTS") when they have a projection source not worth a full database import. Fixed nicks
# (not derived from a database name, since there isn't one) so they stay stable across runs.
# Unlike every database-driven source, fill_source_sheet is NEVER called for one of these (see
# main below) -- these three are the one place in this whole script something manual and
# persistent is supposed to survive a rerun untouched; ensure_raw_source_sheet/
# ensure_source_named_ranges still apply, so a missing one gets its scaffold recreated and its
# Proj<nick>* named ranges stay pointed at it, exactly like a database-driven source.
MANUAL_SOURCES = [
    ("Import 1", "Import 1", "i1"),
    ("Import 2", "Import 2", "i2"),
    ("Import 3", "Import 3", "i3"),
]

# Sheet titles never copied over from the template when bootstrapping RESULT, and pruned from
# RESULT itself if already present (see gsheets_io.open_result_workbook/_prune_dropped_sheets).
# Nothing currently belongs on this list -- every previously-stale/retired/orphaned tab this
# used to name (including, at one point, Import 1/2/3 themselves) has already been removed from
# the template directly rather than merely filtered out of the copy.
DROP_SHEETS = set()


def header_map(ws, row=1):
    """{header text (stripped) -> 1-based column index} for a sheet's header row."""
    out = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=c).value
        if isinstance(v, str) and v.strip():
            out[v.strip()] = c
    return out


def write_row(ws, row, hmap, values: dict):
    """values: {header text -> value}. Skips headers not present in hmap."""
    for h, v in values.items():
        c = hmap.get(h)
        if c is not None and v is not None:
            ws.cell(row=row, column=c, value=v)


def clear_data_rows(ws, first_row, last_row, first_col=1, last_col=None):
    last_col = last_col or ws.max_column
    for r in range(first_row, last_row + 1):
        for c in range(first_col, last_col + 1):
            ws.cell(row=r, column=c).value = None


def clear_namefix(ws):
    """Blanks NameFix's manual-workflow content every run -- kept blank rather than dropping
    the sheet entirely (see DROP_SHEETS) so its "fixnames" named range still exists for other
    tabs' formulas to reference, but with no manual entries: exact-string database names mean
    there's nothing left for a human to alias. IFNA(VLOOKUP(...,fixnames,...),) still resolves
    fine against an empty range -- a lookup that finds nothing is exactly what IFNA is there to
    catch. A2:B<last> is the alias list itself ("DOESN'T MATCH"/"FIXED", header at row 1 kept);
    C1 rightward is a second, separate scratchpad for finding names to add to that list --
    hardcoded to one specific (and no longer active) source sheet and partly built on CONTAINS,
    which was never a real Sheets function -- cleared in full since it has no header worth
    keeping and is just as dead as the list it used to feed."""
    clear_data_rows(ws, 2, ws.max_row, first_col=1, last_col=2)
    clear_data_rows(ws, 1, ws.max_row, first_col=3, last_col=ws.max_column)


# ---------------------------------------------------------------------------
# Phase 0: Database queries -- everything below reads Reference.Players' one canonical
# FullName/PositionCode per player instead of each raw sheet's own name spelling, so every
# tab this script fills uses the exact same string for the same player. That's what makes
# the name-matching formulas elsewhere in the workbook (TeamPOS, VorpAll/PtsVorpAll, etc.)
# actually reliable -- no more nickname/spelling mismatches (e.g. "Matt Boldy" vs. "Matthew
# Boldy") silently dropping a player's position, which is what previously left CValsVorp/
# FanPtsVorp blank for real players. NameFix (a manual alias sheet nothing in code ever read)
# is retired along with the CSV/XLSX-parsing this replaces (aggregate_sources.py) -- its own
# rows are cleared every run by clear_namefix, since the database makes manual aliasing
# unnecessary; only the sheet and its "fixnames" named range are kept around (see DROP_SHEETS).
#
# +/-, split PPG/PPA, raw FOW/FOL (skaters) and GS/GA/SA/SV (goalies) are modeled in the
# database (Projections.SkaterProjections/GoalieProjections) but not every source populates
# every one of them -- write_row() skips any key whose value comes back None, so a source
# that doesn't have a given stat just leaves that cell blank rather than erroring.
# ---------------------------------------------------------------------------

def _skater_projection_rows(cursor, source_name: str) -> dict:
    cursor.execute(
        """
        SELECT p.FullName, p.PositionCode, t.Abbreviation AS Team,
               sp.GamesPlayed, sp.Goals, sp.Assists, sp.Points,
               sp.PowerPlayPoints, sp.ShortHandedPoints, sp.Shots, sp.Hits,
               sp.Blocks, sp.PenaltyMinutes, sp.AverageTOIMinutes,
               sp.PlusMinus, sp.PowerPlayGoals, sp.PowerPlayAssists,
               sp.FaceoffWins, sp.FaceoffLosses
        FROM Projections.SkaterProjections sp
        JOIN Projections.Sources src ON src.SourceID = sp.SourceID
        JOIN Reference.Seasons se ON se.SeasonID = src.SeasonID
        JOIN Reference.Players p ON p.PlayerID = sp.PlayerID
        LEFT JOIN Reference.Teams t ON t.TeamID = sp.TeamID
        WHERE src.SourceName = ? AND se.NHLSeasonID = ?
        """,
        source_name, SEASON_NHL_ID,
    )
    out = {}
    for row in cursor.fetchall():
        out[row.FullName] = {
            "Team": row.Team, "Pos": row.PositionCode, "GP": row.GamesPlayed,
            # dtz.py stores the CSV's season-total 'Total TOI' under this column name
            # (a pre-existing upstream mislabeling, not fixed here) -- SKATER_FIELDS' own
            # TOI/GP division to derive ATOI still works correctly given that.
            "TOI": float(row.AverageTOIMinutes) if row.AverageTOIMinutes is not None else None,
            "G": row.Goals, "A": row.Assists, "PTS": row.Points,
            "PPP": row.PowerPlayPoints, "SHP": row.ShortHandedPoints,
            "HIT": row.Hits, "BLK": row.Blocks, "PIM": row.PenaltyMinutes,
            "SOG": row.Shots,
            "PM": row.PlusMinus, "FOW": row.FaceoffWins, "FOL": row.FaceoffLosses,
            "PPG": row.PowerPlayGoals, "PPA": row.PowerPlayAssists,
        }
    return out


def _goalie_projection_rows(cursor, source_name: str) -> dict:
    cursor.execute(
        """
        SELECT p.FullName, p.PositionCode, t.Abbreviation AS Team,
               gp.GamesPlayed, gp.Wins, gp.Losses, gp.OvertimeLosses,
               gp.GoalsAgainstAverage, gp.SavePercentage, gp.Shutouts,
               gp.GamesStarted, gp.GoalsAgainst, gp.ShotsAgainst, gp.Saves
        FROM Projections.GoalieProjections gp
        JOIN Projections.Sources src ON src.SourceID = gp.SourceID
        JOIN Reference.Seasons se ON se.SeasonID = src.SeasonID
        JOIN Reference.Players p ON p.PlayerID = gp.PlayerID
        LEFT JOIN Reference.Teams t ON t.TeamID = gp.TeamID
        WHERE src.SourceName = ? AND se.NHLSeasonID = ?
        """,
        source_name, SEASON_NHL_ID,
    )
    out = {}
    for row in cursor.fetchall():
        out[row.FullName] = {
            "Team": row.Team, "Pos": row.PositionCode, "GP": row.GamesPlayed,
            "W": row.Wins, "L": row.Losses, "OTL": row.OvertimeLosses,
            "GAA": float(row.GoalsAgainstAverage) if row.GoalsAgainstAverage is not None else None,
            "SVPCT": float(row.SavePercentage) if row.SavePercentage is not None else None,
            "SO": row.Shutouts,
            "GS": row.GamesStarted, "GA": row.GoalsAgainst,
            "SA": row.ShotsAgainst, "SV": row.Saves,
        }
    return out


def _platform_positions(cursor, platform_name: str) -> dict:
    """{FullName -> comma-joined PositionCode string} for one Fantasy platform, e.g.
    'C,LW' for a multi-eligible player."""
    cursor.execute(
        """
        SELECT p.FullName, fp.PositionCode
        FROM Fantasy.PlayerPositions fp
        JOIN Fantasy.Platforms pl ON pl.FantasyPlatformID = fp.FantasyPlatformID
        JOIN Reference.Players p ON p.PlayerID = fp.PlayerID
        JOIN Reference.Seasons se ON se.SeasonID = fp.SeasonID
        WHERE pl.PlatformName = ? AND se.NHLSeasonID = ?
        """,
        platform_name, SEASON_NHL_ID,
    )
    out = {}
    for row in cursor.fetchall():
        out.setdefault(row.FullName, set()).add(row.PositionCode)
    return {name: ",".join(sorted(codes)) for name, codes in out.items()}


def _platform_adp(cursor, platform_name: str) -> dict:
    cursor.execute(
        """
        SELECT p.FullName, fa.ADP
        FROM Fantasy.PlayerADP fa
        JOIN Fantasy.Platforms pl ON pl.FantasyPlatformID = fa.FantasyPlatformID
        JOIN Reference.Players p ON p.PlayerID = fa.PlayerID
        JOIN Reference.Seasons se ON se.SeasonID = fa.SeasonID
        WHERE pl.PlatformName = ? AND se.NHLSeasonID = ?
        """,
        platform_name, SEASON_NHL_ID,
    )
    return {row.FullName: float(row.ADP) for row in cursor.fetchall()}


def _primary_positions(cursor) -> dict:
    """{FullName -> PositionCode}, real on-ice position for every player Reference.Players
    knows about -- the last-resort fallback so a platform with no eligibility data for a
    given player (a deep prospect ESPN doesn't carry, say) still gets *some* position rather
    than a blank that would zero out its VORP/PRNK downstream."""
    cursor.execute("SELECT FullName, PositionCode FROM Reference.Players")
    return {row.FullName: row.PositionCode for row in cursor.fetchall()}


def load_master_data(cursor) -> dict:
    global ACTIVE_SOURCES, DB_EXPORTED_SOURCES, ACTIVE_SOURCE_SHEETS
    db_sources = _active_sources(cursor)
    ACTIVE_SOURCES = db_sources + MANUAL_SOURCES
    DB_EXPORTED_SOURCES = {label for label, _sheet, _nick in ACTIVE_SOURCES}
    ACTIVE_SOURCE_SHEETS = [sheet for _label, sheet, _nick in ACTIVE_SOURCES]

    # {source_name: {"skaters": {...}, "goalies": {...}}} -- database-driven sources only
    # (db_sources, not ACTIVE_SOURCE_SHEETS -- querying Projections.SkaterProjections for
    # "Import 1" would just waste a round trip for a guaranteed-empty result, since it isn't a
    # SourceName that table has ever heard of). A source with no goalie rows at all (e.g. a
    # skater-only export) just gets an empty goalies dict back, same outcome as never querying
    # it, so nothing here needs to know in advance which sources carry which position. main()
    # below is what actually knows to skip fill_source_sheet for MANUAL_SOURCES -- their sheets
    # simply aren't in this dict, so master["sources"][name] would KeyError if it tried.
    sources = {
        name: {
            "skaters": _skater_projection_rows(cursor, name),
            "goalies": _goalie_projection_rows(cursor, name),
        }
        for name, _sheet, _nick in db_sources
    }

    # a name never appears as both skater and goalie across sources in this data; if it did,
    # treat as goalie (rarer, more specific signal) -- same rule aggregate_sources.py used.
    skater_names, goalie_names = set(), set()
    for data in sources.values():
        skater_names |= set(data["skaters"])
        goalie_names |= set(data["goalies"])
    skater_names -= goalie_names

    master = {
        "sources": sources,
        "skater_names": sorted(skater_names), "goalie_names": sorted(goalie_names),
    }

    # Fantasy-PLATFORM (Yahoo/Fantrax/ESPN the apps players draft on, Fantasy.PlayerPositions/
    # PlayerADP) eligibility/ADP -- unrelated to Dom's excluded stat-projection source above
    # (see DB_SOURCE_EXCLUDE) despite the name collision with its former SourceName; feeds
    # Positions!D/E/F and ADPYahoo/ADPFantrax, both kept.
    master["primary_pos"] = _primary_positions(cursor)
    master["yahoo_platform_pos"] = _platform_positions(cursor, "Yahoo")
    master["fantrax_platform_pos"] = _platform_positions(cursor, "Fantrax")
    master["espn_platform_pos"] = _platform_positions(cursor, "ESPN")
    master["yahoo_adp"] = _platform_adp(cursor, "Yahoo")
    master["fantrax_adp"] = _platform_adp(cursor, "Fantrax")
    return master


# ---------------------------------------------------------------------------
# Phase 0.6/1: raw source sheet scaffolding + fill (one sheet per active source, all built the
# same generic way)
#
# The template carries no raw per-source tabs at all now (removed per user instruction --
# fill_source_sheet below overwrites their content completely every run anyway, so a template
# copy was pure leftover clutter), and which sources exist isn't a hardcoded list either (see
# _active_sources) -- so this script has to be fully self-sufficient per source: create the
# sheet if RESULT doesn't have it yet, fill it, and (re)point its Proj<nick>/Names/Cats named
# ranges at it every run, since nothing else provides any of that anymore.
#
# _skater_projection_rows/_goalie_projection_rows already return one uniform dict shape
# regardless of source (the database schema doesn't vary per source, only which fields a given
# source actually populates does) -- SKATER_FIELDS/GOALIE_FIELDS below is the one place that
# shape is named, as {display header -> dict key or a (dict key, transform) pair for the couple
# of fields that need one, e.g. ATOI from TOI/GP}. Both the header row (just the ordered key
# list) and every source's actual data row (built by evaluating this same map) come from it, so
# there's exactly one field list to keep in sync with the database's own columns -- not a
# separate header list per source, and not one hand-kept list duplicating what write_row's
# dict literals already said. (The database itself has no reference table naming these fields'
# short codes -- Projections.Sources only names the source, not its columns -- so this map, not
# a query, is the actual source of truth; a genuinely new stat would still need a line added
# here once, the same as it needing new SkaterProjections/GoalieProjections columns first.)
# ---------------------------------------------------------------------------

SKATER_FIELDS = {
    "Team": "Team", "Pos": "Pos", "GP": "GP",
    "ATOI": lambda s: (s["TOI"] / s["GP"]) if s.get("TOI") and s.get("GP") else None,
    "G": "G", "A": "A", "P": "PTS", "+/-": "PM", "PIM": "PIM",
    "PPG": "PPG", "PPA": "PPA", "PPP": "PPP", "SHP": "SHP", "SOG": "SOG",
    "FOW": "FOW", "FOL": "FOL", "HIT": "HIT", "BLK": "BLK",
}
GOALIE_FIELDS = {
    "Team": "Team", "Pos": "Pos", "GP": "GP", "GS": "GS",
    "W": "W", "L": "L", "OTL": "OTL", "SO": "SO",
    "SV": "SV", "SV%": "SVPCT", "GA": "GA", "GAA": "GAA", "SA": "SA",
}
# Player first, then every field either map can produce, skater fields before goalie-only ones
# (matches this sheet's own row layout: skater rows first, goalie rows after) -- order is
# cosmetic (write_row looks columns up by header text, not position) but keeps a freshly
# created sheet's header row readable.
RAW_SOURCE_HEADERS = (
    ["Player"] + list(SKATER_FIELDS) + [h for h in GOALIE_FIELDS if h not in SKATER_FIELDS]
)
RAW_SOURCE_SCAFFOLD_ROWS = 1000  # matches these tabs' own historical row count


def _row_values(name, row, field_map):
    out = {"Player": name}
    for header, spec in field_map.items():
        out[header] = spec(row) if callable(spec) else row.get(spec)
    return out


def fill_source_sheet(ws, skaters: dict, goalies: dict):
    """Writes one source's skater rows (if any) then goalie rows (if any), using
    SKATER_FIELDS/GOALIE_FIELDS' shared field set for every source uniformly -- write_row
    already skips any header a given row's dict doesn't have a value for, so a skater-only
    source (no GAA/SV%/... at all) or one missing a particular stat just leaves those cells
    blank, exactly as when each source had its own hand-picked header list."""
    hmap = header_map(ws)
    clear_data_rows(ws, 2, ws.max_row, first_col=3)
    r = 2
    for name in sorted(skaters):
        write_row(ws, r, hmap, _row_values(name, skaters[name], SKATER_FIELDS))
        r += 1
    for name in sorted(goalies):
        write_row(ws, r, hmap, _row_values(name, goalies[name], GOALIE_FIELDS))
        r += 1
    return r - 2


def ensure_raw_source_sheet(wb, title):
    """Returns wb[title] -- creating it fresh via CompatWorkbook.add_worksheet (scaffold only,
    no player data; fill_source_sheet writes that immediately after) if RESULT has never had
    this source's sheet before, whether because it's brand new (Projections.Sources just
    started carrying it) or because this is a from-scratch bootstrap (e.g. next season's new
    RESULT workbook). Idempotent and a no-op in the normal case, where the sheet already
    exists."""
    if title in wb.sheetnames:
        return wb[title]
    ws = wb.add_worksheet(title, rows=RAW_SOURCE_SCAFFOLD_ROWS, cols=2 + len(RAW_SOURCE_HEADERS))
    ws.cell(row=1, column=1, value=(
        '=COUNTA(C2:C)&" PLAYERS"&CHAR(10)&COUNTIFS(A2:A,"<>NO MATCH",A2:A,"?*")&" ON MASTER"'
    ))
    ws.cell(row=1, column=2, value=(
        '=SUM(B2:B)&" MATCHED"&char(10)&COUNTIF(B2:B,"?*")&" FIXED"&CHAR(10)&'
        'COUNTIF(A2:A,"NO MATCH")&" NO MATCH"'
    ))
    for i, h in enumerate(RAW_SOURCE_HEADERS):
        ws.cell(row=1, column=3 + i, value=h)
    for r in range(2, RAW_SOURCE_SCAFFOLD_ROWS + 1):
        ws.cell(row=r, column=1, value=f'=IF(C{r}="","",if(B{r}=1,C{r},if(B{r}="","NO MATCH",B{r})))')
        ws.cell(row=r, column=2, value=(
            f'=IFNA(IF(COUNTIF(NamesMasterList,C{r}),1,VLOOKUP(C{r},fixnames,2,FALSE)),)'
        ))
    log.info("%s: created new source sheet (RESULT never had this tab before)", title)
    return ws


def ensure_source_named_ranges(wb, title, nick):
    """(Re)points Proj<nick>/Proj<nick>Names/Proj<nick>Cats at `title` every run --
    AllProjections_S/G's blending formulas (INDIRECT("Proj"&nick) etc, see
    rebuild_all_projections) depend on these existing, but nothing copies them in from the
    template anymore now that it carries no source sheets at all, so this script has to keep
    them correct itself rather than relying on a one-time template-provided copy."""
    last_col = 2 + len(RAW_SOURCE_HEADERS)
    last_row = RAW_SOURCE_SCAFFOLD_ROWS
    set_defined_name(wb, f"Proj{nick}", f"'{title}'!$A$1:${cl(last_col)}${last_row}")
    set_defined_name(wb, f"Proj{nick}Names", f"'{title}'!$A$1:$A${last_row}")
    set_defined_name(wb, f"Proj{nick}Cats", f"'{title}'!$A$1:${cl(last_col)}$1")


# ---------------------------------------------------------------------------
# Phase 2: ADPYahoo / ADPFantrax / ADPother
# ---------------------------------------------------------------------------

def fill_adp_yahoo(ws, yahoo_adp: dict):
    """A:B feeds the ADPYahoo named range (Rankings!D2). (J:M used to also feed
    Positions!D2 via INDEX/MATCH -- now that fill_positions writes Positions!D directly from
    the database, those columns had nothing left reading them and were removed.)"""
    clear_data_rows(ws, 2, ws.max_row)
    r = 2
    for name in sorted(yahoo_adp):
        # only written when a real ADP exists -- VLOOKUP finds a row with a genuinely blank
        # B and returns 0 (not an error), which would otherwise get averaged in as a false
        # "ADP 0 = consensus #1 pick".
        ws.cell(row=r, column=1, value=name)  # A Player
        ws.cell(row=r, column=2, value=yahoo_adp[name])  # B ADP
        r += 1
    return r - 2


def fill_adp_fantrax(ws, doms_adp: dict):
    """A:B feeds the ADPFantrax named range (Rankings!E2). Not read by Positions."""
    clear_data_rows(ws, 2, ws.max_row)
    r = 2
    for name in sorted(doms_adp):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=doms_adp[name])
        r += 1
    return r - 2


def clear_adp_other(ws):
    clear_data_rows(ws, 2, ws.max_row)


# ---------------------------------------------------------------------------
# Phase 3: Positions
# ---------------------------------------------------------------------------

POSITIONS_MANUAL_EXTRA_COL = "T"  # right beside the SKATERS/GOALIES/Included filter columns
POSITIONS_MANUAL_EXTRA_HEADER = "ADD NAMES"
# Short on purpose -- T is a narrow (~80px) column sitting directly against column U's
# "SKATERS" label (a leftover dead filter header, see the SNames/GNames comment on
# _repatch_shifted_positions_refs), so there's no overflow room at all past T's own width.
# The long-form explanation lives in a cell note on T1 instead (set once, live -- this
# function doesn't manage notes) rather than in the header text itself.


def fill_positions(ws, master):
    """D/E/F (Yahoo/Fantrax/ESPN) are written as plain PositionCode values, not formulas --
    each one is already resolved per-platform straight from the database (Fantasy.
    PlayerPositions), falling back to the player's real on-ice PositionCode
    (Reference.Players, present for every player) when a platform has no eligibility row for
    them. This is what actually fixes the blank-VORP/PRNK bug: no more multi-tab INDEX/MATCH
    fallback chain that silently comes up empty when a name doesn't match verbatim. G
    (Fleaflicker) stays blank -- no platform-wide Fleaflicker data exists.

    Column T (POSITIONS_MANUAL_EXTRA_COL) is a standing, never-cleared manual-entry area: when
    a user's own imported spreadsheet has a player the database doesn't know about at all (so
    NameFix's dropdown and MISC3's "NO MATCH" check, both keyed off NamesMasterList, can never
    recognize that name no matter what alias they type), they type the name into T here and it
    gets spliced onto the end of the real database-driven block in column A every run --
    deduped against the database's own names, in whatever order they were typed. This does NOT
    wire the player into any projection/stat pipeline (those all come from the database) --
    it only makes the bare name matchable. NamesMasterList/TeamPOS are resized to cover the
    extended range at the call site, same as AllProj/AllProjCats are for AllProjections_S/G."""
    clear_data_rows(ws, 2, ws.max_row, first_col=1, last_col=7)
    all_names = master["skater_names"] + master["goalie_names"]
    primary_pos = master["primary_pos"]
    yahoo_pos = master["yahoo_platform_pos"]
    fantrax_pos = master["fantrax_platform_pos"]
    espn_pos = master["espn_platform_pos"]
    r = 2
    for name in sorted(all_names):
        ws.cell(row=r, column=1, value=name)  # A Player
        ws.cell(row=r, column=2, value=pick_team(master, name))  # B TEAM
        ws.cell(row=r, column=3, value=(
            f'=CHOOSE(MATCH(Settings!$B$16,$D$1:$G$1,0),D{r},E{r},F{r},G{r})'
        ))
        ws.cell(row=r, column=4, value=yahoo_pos.get(name) or primary_pos.get(name))    # D Yahoo
        ws.cell(row=r, column=5, value=fantrax_pos.get(name) or primary_pos.get(name))  # E Fantrax
        ws.cell(row=r, column=6, value=espn_pos.get(name) or primary_pos.get(name))     # F ESPN
        r += 1

    known = set(all_names)
    extra_col = ci(POSITIONS_MANUAL_EXTRA_COL)
    ws.cell(row=1, column=extra_col, value=POSITIONS_MANUAL_EXTRA_HEADER)
    seen_extra, n_extra = set(), 0
    for er in range(2, ws.max_row + 1):
        name = ws.cell(row=er, column=extra_col).value
        if not isinstance(name, str) or not name.strip():
            continue
        name = name.strip()
        if name in known or name in seen_extra:
            continue
        seen_extra.add(name)
        ws.cell(row=r, column=1, value=name)  # B..G left blank -- no database backing
        r += 1
        n_extra += 1

    return r - 2, all_names, n_extra


def pick_team(master, name):
    for data in master["sources"].values():
        for players in (data["skaters"], data["goalies"]):
            if name in players and players[name].get("Team"):
                return players[name]["Team"]
    return None


# ---------------------------------------------------------------------------
# Phase 4: AllProjections_S / AllProjections_G player list
# ---------------------------------------------------------------------------

def refresh_allprojections_names(ws, names, name_col_letter="A"):
    col = ci(name_col_letter)
    r = 5
    for name in sorted(names):
        ws.cell(row=r, column=col, value=name)
        r += 1
    for row in range(r, ws.max_row + 1):
        ws.cell(row=row, column=col).value = None
    return r - 5


def refresh_export_names(ws, names):
    """Export!A2: (row 2, not 5 -- no frozen header rows here) feeds its own per-player
    INDEX/MATCH-against-AllProj/AllProjNames formulas in B: onward, hardcoded to the skater
    namespace only (no AllGProj equivalent) -- so `names` should be skater_names, matching
    AllProjections_S's own name list. Nothing wrote to this sheet before the template cleanup
    (its player list was just whatever static text happened to be typed in); the demo names
    are gone now, so this keeps it populated without depending on that leftover state."""
    r = 2
    for name in sorted(names):
        ws.cell(row=r, column=1, value=name)
        r += 1
    for row in range(r, ws.max_row + 1):
        ws.cell(row=row, column=1).value = None
    return r - 2


# ---------------------------------------------------------------------------
# Phase 5: Settings source weights
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Phase 6: Player Values - Cats / Player Values - Pts
# ---------------------------------------------------------------------------

# Player Values' row formula isn't one pattern -- it's two, chosen per player: a skater row
# looks stats up in AllProj/AllProjNames/AllProjCats, a goalie row in the AllGProj equivalent
# (see rebuild_all_projections). Rather than hardcode both variants' formula text here (which
# is exactly the "have to change it in the script every run" the template cleanup was meant to
# get away from), each sheet keeps two live reference rows -- a skater example at
# skater_ref_row, a goalie example at goalie_ref_row -- sitting well past the real data block
# (see rebuild_player_values' ref_offset for exactly how far, and why it's not just "the next
# column"; e.g. Player Values - Cats' data ends at AL, so its reference rows live at BG:CR).
# Editing a formula there changes what every real player row gets, with no code change: this
# function reads those two rows once, then row-shifts whichever one matches each player onto
# their real row. The reference rows are seeded from this exact machinery's own previously-
# hardcoded output (see git history), so they start out provably identical to what this sheet
# already produced -- including the already-fixed VAL floor (-9999, not the more obvious -30;
# see the reference row's own G-column formula for why).

_CELL_REF_RE = re.compile(
    r'("(?:[^"]|"")*")'  # 1: a double-quoted string literal, consumed whole and left untouched
    r'|((?:\'[^\']+\'|[A-Za-z_]\w*)!)?(\$?)([A-Z]{1,3})(\$?)(\d+)'  # 2-6: optional sheet prefix, col $, col letters, row $, row number
)


def _shift_row_refs(formula, from_row, to_row):
    """Rewrites every cell reference in `formula` whose row component is unanchored ($-free)
    and equals from_row to instead read to_row, leaving row-anchored references ($2, A$1 --
    the header-row lookups every Player Values formula makes), column letters, sheet-name
    prefixes, and named ranges (which never match this pattern at all -- no named range used
    here ends in digits) untouched. Skips anything inside a double-quoted string literal so a
    stray digit in a text argument can never be mistaken for a row reference."""
    def repl(m):
        if m.group(1) is not None:
            return m.group(1)
        sheet_prefix, col_dollar, col_letters, row_dollar, row_num = (
            m.group(2), m.group(3), m.group(4), m.group(5), m.group(6)
        )
        if row_dollar or int(row_num) != from_row:
            return m.group(0)
        return f"{sheet_prefix or ''}{col_dollar}{col_letters}{row_dollar}{to_row}"
    return _CELL_REF_RE.sub(repl, formula)


def _read_template_row(ws, row, col_offset, last_col):
    """{col_letter: value} for every non-blank cell in [A, last_col_letter] of `row`, shifted
    col_offset columns to the right (where the reference rows actually live)."""
    out = {}
    for c in range(1, last_col + 1):
        v = ws.cell(row=row, column=c + col_offset).value
        if v not in (None, ""):
            out[cl(c)] = v
    return out


def _row_formulas_from_template(template_cells, from_row, to_row):
    """template_cells: a _read_template_row() result. Formula cells (values starting with "=")
    get row-shifted to to_row; anything else (a bare False literal, e.g.) is returned as-is,
    since it doesn't depend on row number. Self-checks each shifted formula by shifting it back
    to from_row and asserting it exactly reproduces the template text -- a shifting bug then
    fails loudly here instead of silently writing a wrong formula to the sheet."""
    out = {}
    for col, val in template_cells.items():
        if isinstance(val, str) and val.startswith("="):
            shifted = _shift_row_refs(val, from_row, to_row)
            if _shift_row_refs(shifted, to_row, from_row) != val:
                raise RuntimeError(
                    f"formula row-shift didn't round-trip for column {col} "
                    f"(from_row={from_row}, to_row={to_row}): {val!r} -> {shifted!r}"
                )
            out[col] = shifted
        else:
            out[col] = val
    return out


def rebuild_player_values(ws, all_names, goalie_names, last_col_letter,
                           skater_ref_row=1, goalie_ref_row=2):
    goalie_set = set(goalie_names)
    last_col = ci(last_col_letter)
    # Reference rows sit well past the real data block -- NOT just one column past last_col.
    # Row 2's own header formula (=TRANSPOSE(CatsUsed) or CatsPtsUsed) declares a spill footprint
    # as wide as its *named range's row count* (misc!D2:D29 / E2:E29, 28 rows each, the fixed
    # 17-skater + 11-goalie category slot count from Settings) regardless of how few categories
    # are actually in use right now -- confirmed live: with last_col=AL(38) a same-sheet offset
    # of just +1 landed the reference rows at AN, which collided with that 28-wide spill's own
    # reach (M through AN) and broke it (#REF! "would overwrite data in AN2"). +20 clears that
    # with real margin on both sheets, not just past the current edge case.
    ref_offset = last_col + 20
    skater_template = _read_template_row(ws, skater_ref_row, ref_offset, last_col)
    goalie_template = _read_template_row(ws, goalie_ref_row, ref_offset, last_col)

    first_row = 3
    last_row = first_row + len(all_names) - 1
    clear_data_rows(ws, first_row, max(ws.max_row, last_row), first_col=1, last_col=last_col)
    r = first_row
    for name in all_names:
        ws.cell(row=r, column=4, value=name)  # D PLAYER
        is_goalie = name in goalie_set
        template = goalie_template if is_goalie else skater_template
        from_row = goalie_ref_row if is_goalie else skater_ref_row
        formulas = _row_formulas_from_template(template, from_row, r)
        for col, val in formulas.items():
            ws.cell(row=r, column=ci(col), value=val)
        r += 1
    return first_row, last_row


# ---------------------------------------------------------------------------
# Phase 5b: SourceCheck full rebuild
#
# The sheet used to be a static template (one hand-authored column per source, each column's
# formulas self-referencing its own column letter, e.g. "'"&F$2&"'" written directly into
# F3:F1078) that this script only ever patched -- fixing a typo, blanking a retired source's
# columns, re-quoting a broken INDIRECT. Every one of those patches existed because deleting
# or adding a column shifts everything after it, but nothing rewrites a *formula's own text*
# to match (openpyxl doesn't do this the way Excel does on a UI column delete, and neither
# would a from-scratch Google Sheets author reliably by hand). rebuild_source_check instead
# regenerates the whole sheet from ACTIVE_SOURCE_SHEETS every run -- exactly as many columns
# as there are active sources, no leftovers, and every self-reference is written using
# whatever column letter that source actually lands on *this* run. Changing the source list
# is now a one-line edit to ACTIVE_SOURCE_SHEETS, not a manual Excel surgery session.
# ---------------------------------------------------------------------------

def rebuild_source_comparison(ws):
    """SourceComparison lets a user pick one skater and one goalie (A2/A16 -- preserved
    as-is, whatever sample name is currently there) and see + average their stats across
    every source. Its per-source row block was hardcoded to 12 rows (one per historical
    source, referencing Settings!S4:S15 by row) the same way SourceCheck's columns used to
    be hardcoded to 12 -- same fix, applied to rows instead of columns: regenerate the block
    to exactly len(ACTIVE_SOURCES) rows every run. Unlike SourceCheck's old self-referencing
    formulas, this sheet's INDEX/MATCH lookups already reference their own row via $B{r} (an
    absolute-column/relative-row ref), so there's no column-letter-style shift bug here --
    the only thing that needed fixing was the row *count*. Rewritten from scratch (not
    row-inserted/deleted) to sidestep openpyxl's lack of formula-text adjustment entirely.
    The category header row (row 1, columns D onward -- SKATERS' own, never moves) and the
    GOALIES section's equivalent (which does move, since the goalie block shifts up as the
    skater block shrinks) are pure static config unrelated to which sources exist -- read
    once and carried over as-is. Row 1 itself is never touched at all."""
    n = len(ACTIVE_SOURCES)
    last_cat_col = ws.max_column

    skater_avg_row = 2
    skater_src_first, skater_src_last = 3, 2 + n
    goalie_header_row = skater_src_last + 1
    goalie_name_row = goalie_header_row + 1
    goalie_src_first, goalie_src_last = goalie_name_row + 1, goalie_name_row + n

    skater_name = ws.cell(row=2, column=1).value
    goalie_header_old_row = next(
        (r for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value == "GOALIES"), None,
    )
    goalie_cats = None
    goalie_name = None
    if goalie_header_old_row is not None:
        # .text, not .value: column C here is an ArrayFormula object (a dead Google-Sheets
        # TRANSPOSE spill) whose internal ref is tied to its original cell -- reassigning
        # that same object to a new coordinate would carry the stale ref along with it, so
        # pull out just the formula text and let it become a plain (still non-functional,
        # same as before) formula at the new position instead.
        goalie_cats = [
            getattr(ws.cell(row=goalie_header_old_row, column=c).value, "text", ws.cell(row=goalie_header_old_row, column=c).value)
            for c in range(1, last_cat_col + 1)
        ]
        goalie_name = ws.cell(row=goalie_header_old_row + 1, column=1).value

    clear_data_rows(ws, 2, ws.max_row, first_col=1, last_col=last_cat_col)
    if ws.max_row > goalie_src_last:
        ws.delete_rows(goalie_src_last + 1, ws.max_row - goalie_src_last)

    ws.cell(row=2, column=1, value=skater_name or "")
    ws.cell(row=2, column=2, value=f'=IFERROR(VLOOKUP(A2,TeamPOS,3,FALSE),"")')
    for r in range(skater_src_first, skater_src_last + 1):
        i = r - skater_src_first
        ws.cell(row=r, column=1, value=f"=Settings!S{4 + i}")
        ws.cell(row=r, column=2, value=f"=Settings!T{4 + i}")
        for c in range(3, last_cat_col + 1):
            letter = cl(c)
            ws.cell(row=r, column=c, value=(
                f'=IFERROR(INDEX(INDIRECT("Proj"&$B{r}),MATCH($A$2,INDIRECT("Proj"&$B{r}&"Names"),0),'
                f'MATCH({letter}$1,INDIRECT("Proj"&$B{r}&"Cats"),0)),)'
            ))
    for c in range(3, last_cat_col + 1):
        letter = cl(c)
        ws.cell(row=skater_avg_row, column=c, value=f"=IFERROR(AVERAGE({letter}{skater_src_first}:{letter}{skater_src_last}),)")

    if goalie_cats is not None:
        for c, v in enumerate(goalie_cats, start=1):
            ws.cell(row=goalie_header_row, column=c, value=v)
    else:
        ws.cell(row=goalie_header_row, column=1, value="GOALIES")
        ws.cell(row=goalie_header_row, column=2, value="POS")
    ws.cell(row=goalie_name_row, column=1, value=goalie_name or "")
    ws.cell(row=goalie_name_row, column=2, value=f'=IFERROR(VLOOKUP(A{goalie_name_row},TeamPOS,3,FALSE),"")')
    for r in range(goalie_src_first, goalie_src_last + 1):
        mirror_row = skater_src_first + (r - goalie_src_first)
        ws.cell(row=r, column=1, value=f"=A{mirror_row}")
        ws.cell(row=r, column=2, value=f"=B{mirror_row}")
        for c in range(3, last_cat_col + 1):
            letter = cl(c)
            ws.cell(row=r, column=c, value=(
                f'=IFERROR(INDEX(INDIRECT("Proj"&$B{r}),MATCH($A${goalie_name_row},INDIRECT("Proj"&$B{r}&"Names"),0),'
                f'MATCH({letter}${goalie_header_row},INDIRECT("Proj"&$B{r}&"Cats"),0)),)'
            ))
    for c in range(3, last_cat_col + 1):
        letter = cl(c)
        ws.cell(row=goalie_name_row, column=c, value=f"=IFERROR(AVERAGE({letter}{goalie_src_first}:{letter}{goalie_src_last}),)")

    return goalie_src_last


def rebuild_source_check(wb, all_names):
    ws = wb["SourceCheck"]
    pos_ws = wb["Positions"]

    # Positions!INCLUDED (a live COUNTIF against AllProjections_S, i.e. "is this player in
    # the final blended pool") is what column B ("MASTER") is meant to check. It used to
    # instead reference Positions' old 'Included' column -- a dead Google-Sheets FILTER()
    # spill that, once exported to Excel, only ever shows one frozen stale name per row, not
    # a real filtered list -- which made MASTER (and everything branching on it below)
    # effectively check against near-random per-row leftovers. Found while rebuilding this
    # sheet; fixed here rather than faithfully reproducing a bug. Located by header text, not
    # a hardcoded letter, so it stays correct if Positions' own layout changes again later.
    included_col = next(
        (c for c in range(1, pos_ws.max_column + 1) if pos_ws.cell(row=1, column=c).value == "INCLUDED"),
        None,
    )
    if included_col is None:
        raise RuntimeError("Positions has no 'INCLUDED' column -- can't rebuild SourceCheck's MASTER check")
    included_letter = cl(included_col)
    # Positions and SourceCheck aren't necessarily row-aligned (Positions is written in
    # sorted(all_names) order; SourceCheck below just uses all_names as given), so this has
    # to be a name-matched INDEX/MATCH lookup, not a same-row reference.

    n_sources = len(ACTIVE_SOURCE_SHEETS)
    total_col = 3 + n_sources
    clear_data_rows(ws, 1, max(ws.max_row, 2 + len(all_names)), first_col=1, last_col=max(ws.max_column, total_col))
    if ws.max_column > total_col:
        ws.delete_cols(total_col + 1, ws.max_column - total_col)

    ws.cell(row=1, column=1, value="last update")
    ws.cell(row=2, column=1, value="PLAYER")
    ws.cell(row=2, column=2, value="MASTER")
    for i, sheet_name in enumerate(ACTIVE_SOURCE_SHEETS):
        ws.cell(row=2, column=3 + i, value=sheet_name)
    ws.cell(row=2, column=total_col, value="TOTAL")

    first_letter, last_letter = cl(3), cl(2 + n_sources)
    r = 3
    for name in all_names:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=(
            f'=IFERROR(INDEX(Positions!${included_letter}:${included_letter},'
            f'MATCH(A{r},Positions!$A:$A,0)),0)'
        ))
        for i, _sheet_name in enumerate(ACTIVE_SOURCE_SHEETS):
            col, letter = 3 + i, cl(3 + i)
            ws.cell(row=r, column=col, value=(
                f'=IF($B{r},COUNTIF(INDIRECT("\'"&{letter}$2&"\'!A:A"),$A{r}),'
                f'COUNTIF(INDIRECT("\'"&{letter}$2&"\'!C:C"),$A{r}))'
            ))
        ws.cell(row=r, column=total_col, value=f"=SUM({first_letter}{r}:{last_letter}{r})")
        r += 1
    return r - 3  # data starts at row 3, not row 2 like the other rebuild_* functions


# ---------------------------------------------------------------------------
# Phase 5c: AllProjections_S/G source-column reduction
#
# Each category (G, A, PIM, ...) occupies a block of columns: one blended (weighted-average)
# value, one z-score, and one raw-value column per source. That per-source sub-range was
# hardcoded to 12 columns (one per historical source -- AGN/BFH/DFO/DtZ/KUB/LX/Cull/SL/YF/i1/
# i2, most now retired) repeated identically across all 17-18 categories. Shrunk here to
# exactly len(ACTIVE_SOURCES), the same source-of-truth list rebuild_source_check and
# rebuild_settings_sources already use.
#
# Two dead-Google-Sheets-spill bugs found and fixed along the way, not faithfully
# reproduced:
#
# 1. Every block's row-4 nick header used to be a single =TRANSPOSE(UsedSRCsABV) spill
#    formula in the block's first source column, with the rest of that block's headers left
#    as plain *frozen* text/0s from whenever this was last live-computed in Google Sheets --
#    e.g. e4:o4 still read 'DtZ','LX','YF',0,0,0,0,0,0,0,0 even after Dailyfaceoff became a
#    4th active source, meaning every block's header row has been silently wrong (not just
#    unused) for a while. Now written directly as plain nick text per column, every run --
#    no spill involved.
#
# 2. UsedSRCsWTs (misc!$N$21:$Y$21), the weight array every block's SUMPRODUCT multiplies
#    by, was itself a dead =TRANSPOSE(N22:N33) spill showing frozen "1"s in every position --
#    meaning every source has actually been getting equal weight the whole time regardless
#    of what Settings' Weight column says (the "Projections Source Confidence" feature
#    described in Settings' own instructions text has never worked once this workbook left
#    Google Sheets). Fixed in rebuild_misc_source_weights, called once before either sheet is
#    rebuilt here.
#
# Row 1/2 (the average/stdev each z-score's STANDARDIZE call needs) is a live Google-Sheets
# FILTER/SORT/ARRAY_CONSTRAIN calculation (top-RosterX-sized slice of that stat, averaged/
# stdev'd), not a static value -- so each block's existing formula is carried over to its new
# column position verbatim (its self-reference is to its OWN column, e.g. H$5:H from a formula
# sitting in column H, so it stays correct wherever the block lands). A category that has
# never had this formula before (found=None, or found but blank -- both categories this
# workbook's history never wired it up for) gets one generated fresh here instead of being
# left blank, using DEFAULT_ZSCORE_ROSTER[sheet_name] as the roster-size named range --
# leaving it blank is what produced #NUM! once real data made STANDARDIZE divide by a blank
# (0) stdev. Row 3 (the VLOOKUP-against-CatsAll/CatWeights used/bonus flags, and the
# per-source category-name chain) is regenerated fresh rather than preserved, since it's a
# generic pattern keyed only by the column's own row-4 content -- unaffected by which physical
# column that ends up being.
# ---------------------------------------------------------------------------

DEFAULT_ZSCORE_ROSTER = {"AllProjections_S": "RosterS", "AllProjections_G": "RosterG"}

# (category representation as written into row 4, kind) per block, in original on-sheet
# order. The representation is usually the plain category code ('G', 'PIM', ...) but a
# handful of blocks instead hold a direct '=Settings!E19'-style cell reference -- some past
# editor's alternate way of naming the same category. Kept exactly as found rather than
# normalized to the usual VLOOKUP(CatsAll,...) style, since changing it risks a typo and
# both forms resolve to the same text either way.
# kind: "first" (the GP column -- source sub-columns, no z-score),
#       "normal" (blended + z-score + source sub-columns),
#       "conditional" (skaters' DEF column -- its blended value is pulled from another
#       block's blended value, conditional on position; z-score, but no source sub-columns
#       of its own).
SKATER_AP_CATEGORIES = [
    ("GP", "first"),
    ("G", "normal"),
    ("A", "normal"),
    ("=Settings!E19", "conditional"),  # DEF -- pulls from the 'P' block, see below
    ("P", "normal"),
    ("+/-", "normal"),
    ("PIM", "normal"),
    ("PPG", "normal"),
    ("PPA", "normal"),
    ("PPP", "normal"),
    ("=Settings!E20", "normal"),  # ATOI
    ("=Settings!E21", "normal"),  # Sy
    ("SHP", "normal"),
    ("SOG", "normal"),
    ("FOW", "normal"),
    ("FOL", "normal"),
    ("HIT", "normal"),
    ("BLK", "normal"),
]
# the conditional (DEF) block's formula references whichever block immediately followed it
# on the original sheet (column AT, the 'P' block) -- captured by category text, not column
# letter, so it stays correct regardless of where blocks land after the rebuild.
SKATER_AP_CONDITIONAL_TARGETS = {"=Settings!E19": "P"}

GOALIE_AP_CATEGORIES = [
    ("GP", "first"),
    ("GS", "normal"),
    ("W", "normal"),
    ("L", "normal"),
    ("GA", "normal"),
    ("GAA", "normal"),
    ("SA", "normal"),
    ("SV", "normal"),
    ("SV%", "normal"),
    ("SO", "normal"),
    ("OTL", "normal"),
    ("=Settings!H15", "normal"),  # Gx
]
GOALIE_AP_CONDITIONAL_TARGETS = {}

_ZSCORE_SELF_REF = re.compile(r"^=if\(([A-Z]+)\$3,")


def _rebind_zscore_formula(formula, new_col_letter):
    """A carried-forward average/stdev formula (see rebuild_all_projections step 1) self-
    references its OWN column (e.g. "=if(H$3, average(...H$5:H...)),...)" sitting in column
    H) -- correct as long as that category's column never moves, but ACTIVE_SOURCES changing
    length (a source added or removed) shifts every later category's column, and the formula
    text carried forward still says the OLD letter. Confirmed empirically: Sheets doesn't
    self-heal this the way it does a real structural column delete -- the stale reference just
    evaluates wrong (or #VALUE!/#NUM! once the column it names holds something else). Rewrites
    every whole-word occurrence of the formula's own self-reference letter to wherever it
    actually landed this run; a formula that already matches its new column, or isn't this
    pattern at all (None, or a category with no carried-forward formula), passes through
    unchanged."""
    if not isinstance(formula, str):
        return formula
    m = _ZSCORE_SELF_REF.match(formula)
    if not m or m.group(1) == new_col_letter:
        return formula
    return re.sub(rf"\b{m.group(1)}\b", new_col_letter, formula)


def rebuild_misc_source_weights(wb):
    """Replaces misc!$N$21:$Y$21 (UsedSRCsWTs, the dead 12-cell =TRANSPOSE(N22:N33) spill
    described above) with a plain, live, len(ACTIVE_SOURCES)-cell array: one
    =IF(Settings!$U{row}=0,,Settings!U{row}) formula per active source's now-compacted
    Settings row (see rebuild_settings_sources), no transpose/spill involved. Also shrinks
    the UsedSRCsWTs defined name to match, since AllProjections_S/G's SUMPRODUCT calls need
    it exactly as wide as the source range they multiply against."""
    ws = wb["misc"]
    n = len(ACTIVE_SOURCES)
    first_row = 4  # rebuild_settings_sources' compacted Settings rows start here
    for i in range(n):
        col = 14 + i  # N, O, P, Q, ...
        settings_row = first_row + i
        ws.cell(row=21, column=col, value=f"=IF(Settings!$U{settings_row}=0,,Settings!U{settings_row})")
    old_last_col = 25  # Y -- the old array's right edge
    new_last_col = 14 + n - 1
    if old_last_col > new_last_col:
        clear_data_rows(ws, 21, 21, first_col=new_last_col + 1, last_col=old_last_col)
    set_defined_name(wb, "UsedSRCsWTs", f"misc!$N$21:${cl(new_last_col)}$21")


def rebuild_all_projections(wb, sheet_name, categories, conditional_targets):
    ws = wb[sheet_name]
    n = len(ACTIVE_SOURCES)
    nicks = [nick for _label, _sheet, nick in ACTIVE_SOURCES]
    old_max_col = ws.max_column
    last_row = ws.max_row

    # 1. locate each category's CURRENT blended column (whatever layout the sheet is in --
    # the old 12-source one on a first run, or this function's own narrow output on a
    # rerun) and capture its row 1/2 cached avg/stdev, the only content that must survive.
    old_avg_stdev = {}
    for cat, _kind in categories:
        found = next(
            (c for c in range(3, ws.max_column + 1) if ws.cell(row=4, column=c).value == cat),
            None,
        )
        old_avg_stdev[cat] = (
            (ws.cell(row=1, column=found).value, ws.cell(row=2, column=found).value) if found else (None, None)
        )

    # 2. compute the new (narrow) column layout.
    new_blended_col = {}
    col = 3
    for cat, kind in categories:
        new_blended_col[cat] = col
        col += {"first": 1 + n, "normal": 2 + n, "conditional": 2}[kind]
    new_last_col = col - 1

    # 3. clear from column C onward (A/B -- player name/pos -- and the row extent are
    # untouched, controlled by refresh_allprojections_names), then trim excess width.
    clear_data_rows(ws, 1, last_row, first_col=3, last_col=old_max_col)
    if old_max_col > new_last_col:
        ws.delete_cols(new_last_col + 1, old_max_col - new_last_col)

    # 4. write every block's headers (rows 1-4) and per-row formulas (rows 5+).
    for cat, kind in categories:
        blended_col = new_blended_col[cat]
        blended_L = cl(blended_col)
        has_zscore = kind in ("normal", "conditional")
        has_sources = kind in ("first", "normal")
        zscore_col = blended_col + 1 if has_zscore else None
        src_first_col = blended_col + (2 if has_zscore else 1) if has_sources else None
        src_last_col = src_first_col + n - 1 if has_sources else None

        ws.cell(row=4, column=blended_col, value=cat)
        if has_zscore:
            ws.cell(row=4, column=zscore_col, value="ZZ")
        if has_sources:
            for i, nick in enumerate(nicks):
                ws.cell(row=4, column=src_first_col + i, value=nick)

        if has_zscore:
            ws.cell(row=3, column=blended_col, value=f"=VLOOKUP({blended_L}$4,CatsAll,2,FALSE())")
            ws.cell(row=3, column=zscore_col, value=f"=VLOOKUP({blended_L}4,CatWeights,3,FALSE())")
        if has_sources:
            for i in range(n):
                sc = src_first_col + i
                ws.cell(row=3, column=sc, value=f"={blended_L}4" if i == 0 else f"={cl(sc - 1)}3")

        a1, a2 = old_avg_stdev[cat]
        a1 = _rebind_zscore_formula(a1, blended_L)
        a2 = _rebind_zscore_formula(a2, blended_L)
        if a1 is None and a2 is None and has_zscore:
            roster = DEFAULT_ZSCORE_ROSTER[sheet_name]
            a1 = (f'=if({blended_L}$3, average(array_constrain(sort(filter({blended_L}$5:{blended_L},'
                  f'{blended_L}$5:{blended_L}<>""),1,false),{roster}+1,1)),"")')
            a2 = (f'=if({blended_L}$3, stdev(array_constrain(sort(filter({blended_L}$5:{blended_L},'
                  f'{blended_L}$5:{blended_L}<>""),1,false),{roster}+1,1)),"")')
        if a1 is not None:
            ws.cell(row=1, column=blended_col).value = a1
        if a2 is not None:
            ws.cell(row=2, column=blended_col).value = a2

        for r in range(5, last_row + 1):
            if kind == "conditional":
                target_L = cl(new_blended_col[conditional_targets[cat]])
                ws.cell(row=r, column=blended_col, value=f'=IFERROR(IF(B{r}="D",{target_L}{r},),)')
            else:
                sL, eL = cl(src_first_col), cl(src_last_col)
                ws.cell(row=r, column=blended_col, value=(
                    f'=IFERROR(SUMPRODUCT({sL}{r}:{eL}{r},UsedSRCsWTs)/'
                    f'SUMIFS(UsedSRCsWTs,{sL}{r}:{eL}{r},"<>"),)'
                ))
                for i in range(n):
                    sc = src_first_col + i
                    scL = cl(sc)
                    ws.cell(row=r, column=sc, value=(
                        f'=IF({scL}$4="",,IFERROR(INDEX(INDIRECT("Proj"&{scL}$4),'
                        f'MATCH($A{r},INDIRECT("Proj"&{scL}$4&"Names"),0),'
                        f'MATCH({scL}$3,INDIRECT("Proj"&{scL}$4&"Cats"),0)),))'
                    ))
            if has_zscore:
                zL = cl(zscore_col)
                ws.cell(row=r, column=zscore_col, value=(
                    f'=IF(OR({blended_L}{r}="",{blended_L}$3=FALSE()),"",IF({zL}$3,'
                    f'MAX(0,STANDARDIZE({blended_L}{r},{blended_L}$1,{blended_L}$2)),'
                    f'STANDARDIZE({blended_L}{r},{blended_L}$1,{blended_L}$2)))'
                ))

    return new_last_col


# ---------------------------------------------------------------------------
# Phase 5d: Schedule Info rebuild
#
# One column per calendar day of the season, one row per team, each cell either blank (no
# game), the opponent's abbreviation (home game), or "@"+opponent (away game) -- was static
# data pasted in from the 2025-26 season (Oct 2025 - Apr 2026) and never refreshed, now a
# full year stale relative to this workbook's actual 2026-27 season. Regenerated here from
# Reference.Schedule. OffNights/PlayoffGames (rows 37-68, read by Player Values - Cats/Pts'
# K/L columns via those exact defined names) keep the same row/column shape -- one row per
# team, same order as rows 2-33 -- so the defined names themselves don't need touching, only
# their formulas' hardcoded last-column reference (originally GK, wherever the new season's
# day count actually puts it).
# ---------------------------------------------------------------------------

def rebuild_schedule_info(wb, cursor, season_nhl_id):
    ws = wb["Schedule Info"]
    cursor.execute(
        """
        SELECT s.GameDate, ht.Abbreviation AS Home, at.Abbreviation AS Away
        FROM Reference.Schedule s
        JOIN Reference.Seasons se ON se.SeasonID = s.SeasonID
        JOIN Reference.Teams ht ON ht.TeamID = s.HomeTeamID
        JOIN Reference.Teams at ON at.TeamID = s.AwayTeamID
        WHERE se.NHLSeasonID = ? AND s.GameType = '2'
        ORDER BY s.GameDate
        """,
        season_nhl_id,
    )
    games = cursor.fetchall()
    if not games:
        raise RuntimeError(f"Reference.Schedule has no games for season {season_nhl_id}")

    min_date = min(g.GameDate for g in games)
    max_date = max(g.GameDate for g in games)
    n_days = (max_date - min_date).days + 1
    teams = sorted({g.Home for g in games} | {g.Away for g in games})
    team_row = {team: 2 + i for i, team in enumerate(teams)}
    last_col = 1 + n_days
    last_L = cl(last_col)

    old_max_row, old_max_col = ws.max_row, ws.max_column
    clear_data_rows(ws, 1, old_max_row, first_col=1, last_col=old_max_col)
    if old_max_col > last_col:
        ws.delete_cols(last_col + 1, old_max_col - last_col)
    new_max_row = 36 + len(teams)
    if old_max_row > new_max_row:
        ws.delete_rows(new_max_row + 1, old_max_row - new_max_row)

    ws.cell(row=1, column=1, value="team")
    for d in range(n_days):
        ws.cell(row=1, column=2 + d, value=min_date + timedelta(days=d))
    for team, r in team_row.items():
        ws.cell(row=r, column=1, value=team)
    for g in games:
        col = 2 + (g.GameDate - min_date).days
        ws.cell(row=team_row[g.Home], column=col, value=g.Away)
        ws.cell(row=team_row[g.Away], column=col, value=f"@{g.Home}")

    ws.cell(row=34, column=1, value="games")
    for d in range(n_days):
        col = 2 + d
        L = cl(col)
        ws.cell(row=34, column=col, value=f"=COUNTA({L}2:{L}33)/2")

    ws.cell(row=36, column=1, value="off night games")
    ws.cell(row=36, column=5, value="playoffs")
    ws.cell(row=36, column=8, value="All")
    ws.cell(row=36, column=9, value="Off")

    for i, team in enumerate(teams):
        r = 37 + i
        src_row = team_row[team]
        rng = f"$B{src_row}:${last_L}{src_row}"
        ws.cell(row=r, column=1, value=team)
        ws.cell(row=r, column=2, value=(
            f'=COUNTIFS({rng},"?*",$B$34:${last_L}$34,"<=8",$B$1:${last_L}$1,"<="&$F$38)'
        ))
        if i == 0:
            ws.cell(row=r, column=5, value="start")
            ws.cell(row=r, column=6, value="=Settings!C19")
        elif i == 1:
            ws.cell(row=r, column=5, value="end")
            ws.cell(row=r, column=6, value="=Settings!C20")
        ws.cell(row=r, column=7, value=team)
        ws.cell(row=r, column=8, value=(
            f'=COUNTIFS({rng},"?*",$B$1:${last_L}$1,">="&$F$37,$B$1:${last_L}$1,"<="&$F$38)'
        ))
        ws.cell(row=r, column=9, value=(
            f'=COUNTIFS({rng},"?*",$B$1:${last_L}$1,">="&$F$37,$B$1:${last_L}$1,"<="&$F$38,'
            f'$B$34:${last_L}$34,"<=8")'
        ))
        ws.cell(row=r, column=10, value=f'=H{r}&" ("&I{r}&")"')

    # OffNights/PlayoffGames stay one row per team starting at row 37 -- resize the defined
    # names themselves (not just leave them at a hardcoded $68) so a future team-count change
    # (expansion) doesn't silently leave rows outside the named range.
    set_defined_name(wb, "OffNights", f"'Schedule Info'!$A$37:$B${new_max_row}")
    set_defined_name(wb, "PlayoffGames", f"'Schedule Info'!$G$37:$J${new_max_row}")

    return {"n_days": n_days, "n_teams": len(teams), "min_date": min_date, "max_date": max_date}


def rebuild_settings_sources(ws):
    """Regenerates Settings' Source/Nick/Weight table (S:U, rows 4-15) from ACTIVE_SOURCES
    (itself database-derived every run, see _active_sources -- not a fixed list), compacted
    into rows 4..3+len(ACTIVE_SOURCES) with weight 1, every row after that cleared -- rather
    than leaving whatever's currently active scattered across old rows, interspersed with
    blanked-out entries for sources that have since stopped being imported.

    Only clears/writes cell VALUES in S:U, never deletes rows: Settings' rows 4-15 also hold
    unrelated category-weight data in columns E:J at the same row numbers, which delete_rows
    would destroy. That means rows past the active count still exist and still look "blank"
    in S:U -- full-width removal isn't possible without moving the table somewhere with no
    unrelated data at those rows, which wasn't judged worth the disruption.

    misc!$M$22:$M$33/$N$21:$Y$21 (UsedSRCsABV/UsedSRCsWTs) read Settings!T4:U15 positionally
    -- each of their 12 slots maps to one specific Settings row -- because that's how many
    columns AllProjections_S/G's own SUMPRODUCT hardcodes (D:O, one per historical source).
    That 12-slot span can't shrink without also restructuring AllProjections_S/G's column
    layout, so misc keeps showing 8 blank positions after this compaction; but since those
    formulas already just read whatever's actually in each row, compacting the real rows to
    the front (so they land in misc's first 4 slots) needs no changes there.

    Each Source cell can also carry a rich-text hyperlink (the source's own URL) in the
    template -- replacing .value here (a plain values.update, not a copy) drops that
    formatting as a side effect, same as it would if typed over by hand."""
    first_row = 4
    last_row = 15
    for i, (label, _sheet, nick) in enumerate(ACTIVE_SOURCES):
        row = first_row + i
        ws.cell(row=row, column=19, value=label)  # S
        ws.cell(row=row, column=20, value=nick)    # T
        ws.cell(row=row, column=21, value=1)       # U
    for row in range(first_row + len(ACTIVE_SOURCES), last_row + 1):
        ws.cell(row=row, column=19).value = None
        ws.cell(row=row, column=20).value = None
        ws.cell(row=row, column=21).value = None
    return len(ACTIVE_SOURCES)


# ---------------------------------------------------------------------------
# Phase 7: Rankings (static player list; B:I formulas already fill down and
# stay untouched -- Excel-native, driven entirely by name-keyed VLOOKUPs)
# ---------------------------------------------------------------------------

def refresh_rankings(ws, all_names):
    clear_data_rows(ws, 2, max(ws.max_row, 2 + len(all_names)), first_col=1, last_col=1)
    r = 2
    for name in all_names:
        ws.cell(row=r, column=1, value=name)
        r += 1
    # Column E's IFERROR fallback is bare (,) not (,"") unlike D/F -- evaluates to 0, which
    # then pollutes column C's AVERAGE(D:F) for anyone missing Fantrax ADP specifically.
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
        for cell in row:
            v = cell.value
            text = getattr(v, "text", v)
            if isinstance(text, str) and text.rstrip().endswith(",)"):
                cell.value = text.rstrip()[:-2] + ',"")'
    return r - 2


# ---------------------------------------------------------------------------
# Phase 8: CValsVorp / FanPtsVorp (VORP calc)
#
# The original used dead FILTER/SORT to build 3 separate position-split, value-sorted
# sub-tables (forwards/D/G), each with its own replacement-level threshold and a running
# rank counter, then dispatched into them by name. That FILTER/SORT half is rebuilt here
# using only pre-2007 functions (SUMPRODUCT+LARGE for the replacement-level threshold,
# SUMPRODUCT tie-broken counting for rank) instead of FILTER/SORTBY/LET, which turned out
# to need Excel's undocumented _xlfn. prefix to work at all when written via openpyxl (see
# fix_ifna_globally) -- rather than chase that per function, everything here sticks to
# functions that have been in Excel since 2007. This also drops the original's per-sub-
# position (C/LW/RW) multi-credit VORP logic in favor of one threshold per group (forward/
# defense/goalie) -- a deliberate simplification, not a fidelity gap: nothing outside this
# sheet ever depended on the sub-position credit, only on VorpAll's VORP/PRNK columns.
# ---------------------------------------------------------------------------

def _group_rank_sumproduct(value_range, cond, myval_ref, row_range, r):
    return (
        f'SUMPRODUCT(({cond})*({value_range}>{myval_ref}))'
        f'+SUMPRODUCT(({cond})*({value_range}={myval_ref})*(ROW({row_range})<{r}))+1'
    )


def rebuild_vorp(ws, source_sheet, last_row, roster_f_name, roster_d_name, roster_g_name):
    """source_sheet: 'Player Values - Cats' or 'Player Values - Pts'. Writes VorpAll's
    A:F (mirrors source D:G, then VORP/PRNK) at CValsVorp!A3:F<last_row>, thresholds at
    H1:J1. Same layout used for both CValsVorp and FanPtsVorp -- caller passes the sheet."""
    old_max_col = ws.max_column
    clear_data_rows(ws, 1, max(ws.max_row, last_row), first_col=1, last_col=old_max_col)
    ws.cell(row=2, column=1, value="PLAYER")
    ws.cell(row=2, column=2, value="TEAM")
    ws.cell(row=2, column=3, value="POS")
    ws.cell(row=2, column=4, value="VAL")
    ws.cell(row=2, column=5, value="VORP")
    ws.cell(row=2, column=6, value="PRNK")

    crange = f"$C$3:$C${last_row}"
    drange = f"$D$3:$D${last_row}"
    fwd_cond = f'({crange}<>"D")*({crange}<>"G")'
    def_cond = f'{crange}="D"'
    gk_cond = f'{crange}="G"'

    ws.cell(row=1, column=8, value=(
        f'=SUMPRODUCT(LARGE(({fwd_cond})*{drange}+(1-({fwd_cond}))*-999999,'
        f'MIN(SUMPRODUCT(({fwd_cond})*1),{roster_f_name}+1)))'
    ))
    ws.cell(row=1, column=9, value=(
        f'=SUMPRODUCT(LARGE(({def_cond})*{drange}+(1-({def_cond}))*-999999,'
        f'MIN(SUMPRODUCT(({def_cond})*1),{roster_d_name}+1)))'
    ))
    ws.cell(row=1, column=10, value=(
        f'=SUMPRODUCT(LARGE(({gk_cond})*{drange}+(1-({gk_cond}))*-999999,'
        f'MIN(SUMPRODUCT(({gk_cond})*1),70,{roster_g_name}+1)))'
    ))

    for r in range(3, last_row + 1):
        ws.cell(row=r, column=1, value=f"='{source_sheet}'!D{r}")
        ws.cell(row=r, column=2, value=f"='{source_sheet}'!E{r}")
        ws.cell(row=r, column=3, value=f"='{source_sheet}'!F{r}")
        ws.cell(row=r, column=4, value=f"='{source_sheet}'!G{r}")
        ws.cell(row=r, column=5, value=(
            f'=IF($C{r}="","",IF($C{r}="D",$D{r}-$I$1,IF($C{r}="G",$D{r}-$J$1,$D{r}-$H$1)))'
        ))
        fwd_rank = _group_rank_sumproduct(drange, fwd_cond, f"$D{r}", f"$D$3:$D${last_row}", r)
        def_rank = _group_rank_sumproduct(drange, def_cond, f"$D{r}", f"$D$3:$D${last_row}", r)
        gk_rank = _group_rank_sumproduct(drange, gk_cond, f"$D{r}", f"$D$3:$D${last_row}", r)
        ws.cell(row=r, column=6, value=(
            f'=IF($C{r}="","",IF($C{r}="D","D"&({def_rank}),IF($C{r}="G","G"&({gk_rank}),'
            f'"F"&({fwd_rank}))))'
        ))


# ---------------------------------------------------------------------------
# Phase 9: CleanCat / CleanPts (feeds Cheat Sheet directly)
#
# Hidden block (A..H for Cats / A..I for Pts) mirrors Player Values row-for-row (already
# Excel-native in the original -- kept as-is, just extended to the new row count). A
# composite rank key ("F0001"/"D0023"/"G0005") is added right after it, then 3 display
# blocks (forwards, defense, goalies) retrieve the k-th player per group via a single
# INDEX/MATCH against that key -- no FILTER/SORT/ARRAY_CONSTRAIN needed.
# ---------------------------------------------------------------------------

def _composite_rank_key(gcell, vcell, group_range, val_range, r):
    fwd_cond = f'({group_range}<>"D")*({group_range}<>"G")'
    def_cond = f'{group_range}="D"'
    gk_cond = f'{group_range}="G"'
    fwd_rank = _group_rank_sumproduct(val_range, fwd_cond, vcell, val_range, r)
    def_rank = _group_rank_sumproduct(val_range, def_cond, vcell, val_range, r)
    gk_rank = _group_rank_sumproduct(val_range, gk_cond, vcell, val_range, r)
    return (
        f'=IF({gcell}="D","D"&TEXT({def_rank},"0000"),'
        f'IF({gcell}="G","G"&TEXT({gk_rank},"0000"),'
        f'"F"&TEXT({fwd_rank},"0000")))'
    )


def rebuild_clean_cat(ws, last_pv_row):
    """CleanCat. Hidden block A:H (H=POS/group key), key col I. Display: J:P forwards,
    R:W defense (6 cols, no G/POG), Z:AE goalies (6 cols)."""
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    src = "Player Values - Cats"
    first_hidden, last_hidden = 4, 4 + (last_pv_row - 3)
    ws.cell(row=2, column=1, value="CATEGORIES")
    ws.cell(row=2, column=11, value="FORWARDS")
    header_src = {"A": "A2", "B": "D2", "C": "G2", "D": "I2", "E": "H2", "F": "L2", "G": "K2"}
    for col, cell in header_src.items():
        ws.cell(row=3, column=ci(col), value=f"='{src}'!{cell}")
    ws.cell(row=3, column=8, value="POS")

    val_range = f"$C${first_hidden}:$C${last_hidden}"
    group_range = f"$H${first_hidden}:$H${last_hidden}"
    for r in range(first_hidden, last_hidden + 1):
        pv = r - 1
        ws.cell(row=r, column=1, value=(
            f'=if(\'{src}\'!A{pv}="","-",ceiling(\'{src}\'!A{pv}/TEAMS,1)&"-"&'
            f'text(if(mod(\'{src}\'!A{pv},TEAMS)=0,TEAMS,if(mod(\'{src}\'!A{pv},TEAMS)<1,1,'
            f'mod(\'{src}\'!A{pv},TEAMS))),"00"))'
        ))
        ws.cell(row=r, column=2, value=f"='{src}'!D{pv}&\" - \"&'{src}'!E{pv}&\" - \"&'{src}'!F{pv}")
        ws.cell(row=r, column=3, value=f"='{src}'!G{pv}")
        ws.cell(row=r, column=4, value=f"='{src}'!I{pv}")
        ws.cell(row=r, column=5, value=f"='{src}'!H{pv}")
        ws.cell(row=r, column=6, value=f"='{src}'!L{pv}")
        ws.cell(row=r, column=7, value=f"='{src}'!K{pv}")
        ws.cell(row=r, column=8, value=f"='{src}'!F{pv}")
        ws.cell(row=r, column=9, value=_composite_rank_key(f"$H{r}", f"$C{r}", group_range, val_range, r))

    key_range = f"$I${first_hidden}:$I${last_hidden}"

    def display_block(anchor_col_letter, n_rows, prefix, src_cols):
        ac = ci(anchor_col_letter)
        for j, sc in enumerate(src_cols):
            ws.cell(row=3, column=ac + j, value=f"={cl(sc)}3")
        for k in range(1, n_rows + 1):
            r = 3 + k
            for j, sc in enumerate(src_cols):
                sl = cl(sc)
                ws.cell(row=r, column=ac + j, value=(
                    f'=IFERROR(INDEX(${sl}${first_hidden}:${sl}${last_hidden},'
                    f'MATCH("{prefix}"&TEXT({k},"0000"),{key_range},0)),"")'
                ))

    display_block("J", 600, "F", [1, 2, 3, 4, 5, 6, 7])
    display_block("R", 260, "D", [1, 2, 3, 4, 5, 6])
    display_block("Z", 100, "G", [1, 2, 3, 4, 5, 6])
    return first_hidden, last_hidden


def rebuild_clean_pts(ws, last_pv_row):
    """CleanPts. Hidden block A:I (I=POS/group key), key col J. Display: K:R forwards,
    T:Y defense, AC:AH goalies (6 cols each for D/G, 8 for forwards)."""
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    src = "Player Values - Pts"
    first_hidden, last_hidden = 4, 4 + (last_pv_row - 3)
    ws.cell(row=2, column=1, value="POINTS")
    ws.cell(row=2, column=12, value="FORWARDS")
    header_src = {"A": "A2", "B": "D2", "C": "G2", "D": "H2", "E": "J2", "F": "I2", "G": "M2", "H": "L2"}
    for col, cell in header_src.items():
        ws.cell(row=3, column=ci(col), value=f"='{src}'!{cell}")
    ws.cell(row=3, column=9, value="POS")

    val_range = f"$C${first_hidden}:$C${last_hidden}"
    group_range = f"$I${first_hidden}:$I${last_hidden}"
    for r in range(first_hidden, last_hidden + 1):
        pv = r - 1
        ws.cell(row=r, column=1, value=(
            f'=if(\'{src}\'!A{pv}="","-",ceiling(\'{src}\'!A{pv}/TEAMS,1)&"-"&'
            f'text(if(mod(\'{src}\'!A{pv},TEAMS)=0,TEAMS,if(mod(\'{src}\'!A{pv},TEAMS)<1,1,'
            f'mod(\'{src}\'!A{pv},TEAMS))),"00"))'
        ))
        ws.cell(row=r, column=2, value=f"='{src}'!D{pv}&\" - \"&'{src}'!E{pv}&\" - \"&'{src}'!F{pv}")
        ws.cell(row=r, column=3, value=f"='{src}'!G{pv}")
        ws.cell(row=r, column=4, value=f"='{src}'!H{pv}")
        ws.cell(row=r, column=5, value=f"='{src}'!J{pv}")
        ws.cell(row=r, column=6, value=f"='{src}'!I{pv}")
        ws.cell(row=r, column=7, value=f"='{src}'!M{pv}")
        ws.cell(row=r, column=8, value=f"='{src}'!L{pv}")
        ws.cell(row=r, column=9, value=f"='{src}'!F{pv}")
        ws.cell(row=r, column=10, value=_composite_rank_key(f"$I{r}", f"$C{r}", group_range, val_range, r))

    key_range = f"$J${first_hidden}:$J${last_hidden}"

    def display_block(anchor_col_letter, n_rows, prefix, src_cols):
        ac = ci(anchor_col_letter)
        for j, sc in enumerate(src_cols):
            ws.cell(row=3, column=ac + j, value=f"={cl(sc)}3")
        for k in range(1, n_rows + 1):
            r = 3 + k
            for j, sc in enumerate(src_cols):
                sl = cl(sc)
                ws.cell(row=r, column=ac + j, value=(
                    f'=IFERROR(INDEX(${sl}${first_hidden}:${sl}${last_hidden},'
                    f'MATCH("{prefix}"&TEXT({k},"0000"),{key_range},0)),"")'
                ))

    display_block("K", 600, "F", [1, 2, 3, 4, 5, 6, 7, 8])
    display_block("T", 260, "D", [1, 2, 3, 4, 5, 6, 7])
    display_block("AC", 100, "G", [1, 2, 3, 4, 5, 6, 7])
    return first_hidden, last_hidden


# ---------------------------------------------------------------------------
# Phase 10: Available - Cats / Pts
#
# 7 blocks each: overall top-30 by ADP (undrafted only), then top-7-by-VAL(or FanPts) for
# each of C/RW/LW/multi-eligible/D/G. Rank helper columns live on Player Values itself
# (one row per player, same alignment as the sheet already uses everywhere else), each
# assigning a unique tie-broken rank within its own undrafted+position group; the display
# blocks here just do IFERROR(INDEX(...,MATCH(k,rank_range,0)),"") -- no QUERY/FILTER/SORT.
# ---------------------------------------------------------------------------

def add_rank_helpers(pv_ws, last_pv_row, first_helper_col_letter, value_col_letter):
    """Writes 7 rank-helper columns onto the Player Values sheet itself (pv_ws), rows
    3..last_pv_row:
      [0] rank by ADP asc among undrafted players with a real ADP (for the overall block)
      [1..6] rank by value_col_letter desc among undrafted players eligible at C/RW/LW/
             multi-position/D/G, in that order
    Returns the list of 7 column letters, in that order."""
    first_col = ci(first_helper_col_letter)
    cols = [cl(first_col + i) for i in range(7)]
    first_pv, last_pv = 3, last_pv_row
    a_r = f"$A${first_pv}:$A${last_pv}"
    f_r = f"$F${first_pv}:$F${last_pv}"
    v_r = f"${value_col_letter}${first_pv}:${value_col_letter}${last_pv}"
    c_r = f"$C${first_pv}:$C${last_pv}"
    notaken = f"({c_r}<>TRUE)"

    pos_test_range = {
        1: f'ISNUMBER(FIND("C",{f_r}))',
        2: f'ISNUMBER(FIND("R",{f_r}))',
        3: f'ISNUMBER(FIND("L",{f_r}))',
        4: f'ISNUMBER(FIND(",",{f_r}))',
        5: f'({f_r}="D")',
        6: f'({f_r}="G")',
    }
    pos_test_cell = {
        1: lambda r: f'ISNUMBER(FIND("C",$F{r}))',
        2: lambda r: f'ISNUMBER(FIND("R",$F{r}))',
        3: lambda r: f'ISNUMBER(FIND("L",$F{r}))',
        4: lambda r: f'ISNUMBER(FIND(",",$F{r}))',
        5: lambda r: f'($F{r}="D")',
        6: lambda r: f'($F{r}="G")',
    }

    for r in range(first_pv, last_pv + 1):
        # Rankings!E2's IFERROR fallback is bare (,) not (,"") -- evaluates to 0, not blank,
        # for players missing Fantrax ADP -- so also exclude literal 0 here, not just "".
        cond0 = f'{notaken}*({a_r}<>"")*({a_r}<>0)'
        rank0 = (
            f'SUMPRODUCT(({cond0})*({a_r}<$A{r}))'
            f'+SUMPRODUCT(({cond0})*({a_r}=$A{r})*(ROW({a_r})<{r}))+1'
        )
        pv_ws.cell(row=r, column=first_col, value=(
            f'=IF($C{r}=TRUE,"",IF($A{r}="","",{rank0}))'
        ))
        for i in range(1, 7):
            cond = f'{notaken}*({pos_test_range[i]})'
            rank = _group_rank_sumproduct(v_r, cond, f"${value_col_letter}{r}", v_r, r)
            pv_ws.cell(row=r, column=first_col + i, value=(
                f'=IF($C{r}=TRUE,"",IF(NOT({pos_test_cell[i](r)}),"",{rank}))'
            ))
    return cols


def _write_available_block(ws, anchor_row, anchor_col_letter, n_rows, rank_col_letter,
                            pv_sheet, pv_first, pv_last, out_pv_cols, header_labels):
    ac = ci(anchor_col_letter)
    for j, label in enumerate(header_labels):
        ws.cell(row=anchor_row, column=ac + j, value=label)
    rank_range = f"'{pv_sheet}'!${rank_col_letter}${pv_first}:${rank_col_letter}${pv_last}"
    for k in range(1, n_rows + 1):
        r = anchor_row + k
        for j, pv_col in enumerate(out_pv_cols):
            pv_range = f"'{pv_sheet}'!${pv_col}${pv_first}:${pv_col}${pv_last}"
            ws.cell(row=r, column=ac + j, value=(
                f'=IFERROR(INDEX({pv_range},MATCH({k},{rank_range},0)),"")'
            ))


def rebuild_available(ws, pv_ws, pv_sheet_name, last_pv_row, value_col_letter, vorp_col_letter):
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    clear_data_rows(ws, 1, max(ws.max_row, 35), first_col=1, last_col=15)
    pv_first, pv_last = 3, last_pv_row
    rank_cols = add_rank_helpers(pv_ws, last_pv_row, "AN", value_col_letter)

    _write_available_block(
        ws, 1, "A", 30, rank_cols[0], pv_sheet_name, pv_first, pv_last,
        ["A", "D", "F"], ["ADP", "PLAYER", "POS"])

    pos_blocks = [
        (3, "E", 1, "C"), (3, "K", 2, "RW"),
        (13, "E", 3, "LW"), (13, "K", 4, "MULTI"),
        (23, "E", 5, "D"), (23, "K", 6, "G"),
    ]
    out_cols = ["D", "F", value_col_letter, vorp_col_letter, "A"]
    headers = ["PLAYER", "POS", "VAL", "VORP", "ADP"]
    for anchor_row, anchor_col, rank_idx, label in pos_blocks:
        hdr = list(headers)
        hdr[0] = f"{label} PLAYER"
        _write_available_block(
            ws, anchor_row, anchor_col, 7, rank_cols[rank_idx], pv_sheet_name,
            pv_first, pv_last, out_cols, hdr)

    # values.update (what wb.save() sends) never touches a cell's format, so these formula
    # writes silently inherit whatever format each cell already had -- which turned out to be
    # none at all here (confirmed live: VAL/VORP showing 10+ raw decimal places instead of the
    # template's "0.00"), however that happened over this workbook's history. Reapplied
    # explicitly every run rather than trusted to already be right. Row range covers all 3
    # position-block row-bands (4-10/14-20/24-30) in one shot -- the gaps between them (11-13,
    # 21-23) get formatted too, harmlessly, since nothing occupies those rows in this sheet.
    ws.set_number_format(2, ci("A"), 31, ci("A"), "0.0")
    ws.set_number_format(4, ci("G"), 30, ci("H"), "0.00")
    ws.set_number_format(4, ci("I"), 30, ci("I"), "0.0")
    ws.set_number_format(4, ci("M"), 30, ci("N"), "0.00")
    ws.set_number_format(4, ci("O"), 30, ci("O"), "0.0")


if __name__ == "__main__":
    log.info("Loading source data from the database...")
    conn = nhl_db.connect()
    cursor = conn.cursor()
    master = load_master_data(cursor)
    conn.close()
    log.info("skaters=%d goalies=%d", len(master["skater_names"]), len(master["goalie_names"]))

    log.info("Opening result spreadsheet (bootstrapping from the template if needed)...")
    wb = gsheets_io.open_result_workbook(
        SOURCE_SPREADSHEET_ID, RESULT_SPREADSHEET_ID, DROP_SHEETS, CREDENTIALS_PATH,
    )

    clear_namefix(wb["NameFix"])
    log.info("NameFix: cleared manual alias rows (fixnames kept, unused now that names come from the database)")

    sched_conn = nhl_db.connect()
    sched_info = rebuild_schedule_info(wb, sched_conn.cursor(), SEASON_NHL_ID)
    sched_conn.close()
    log.info(
        "Schedule Info: rebuilt, %d team(s), %d day(s) (%s - %s)",
        sched_info["n_teams"], sched_info["n_days"], sched_info["min_date"], sched_info["max_date"],
    )

    # Settings!C19:C20 (Playoffs Schedule Start/End Date) is a fantasy-league-specific
    # window into the season the workbook can't derive on its own -- was stuck at the old
    # 2025-26 season's dates (2026-03-16 - 2026-04-05), which no longer even overlaps the
    # new season's date range at all, so OffNights/PlayoffGames' COUNTIFS would find zero
    # matching dates and silently show all zeros. Defaulted here to the new season's last 3
    # weeks; adjust in Settings if your league's actual playoff weeks differ.
    settings_ws = wb["Settings"]
    playoff_end = sched_info["max_date"]
    playoff_start = playoff_end - timedelta(weeks=3)
    settings_ws.cell(row=19, column=3, value=playoff_start)
    settings_ws.cell(row=20, column=3, value=playoff_end)
    log.info("Settings: defaulted Playoffs Schedule to %s - %s (last 3 weeks of season)", playoff_start, playoff_end)

    for name, _sheet, nick in ACTIVE_SOURCES:
        ws = ensure_raw_source_sheet(wb, name)
        if name in master["sources"]:
            n = fill_source_sheet(ws, master["sources"][name]["skaters"], master["sources"][name]["goalies"])
            log.info("%s: %d rows", name, n)
        else:
            log.info("%s: manual source, left untouched", name)
        ensure_source_named_ranges(wb, name, nick)

    n = fill_adp_yahoo(wb["ADPYahoo"], master["yahoo_adp"])
    log.info("ADPYahoo: %d rows", n)

    n = fill_adp_fantrax(wb["ADPFantrax"], master["fantrax_adp"])
    log.info("ADPFantrax: %d rows", n)
    set_defined_name(wb, "ADPFantrax", f"ADPFantrax!$A$1:$B${n + 1}")

    clear_adp_other(wb["ADPother"])
    log.info("ADPother: cleared (no Fleaflicker source)")

    n, all_names, n_extra = fill_positions(wb["Positions"], master)
    log.info("Positions: %d rows (%d from the database + %d manually-added extra name(s))",
              n, len(all_names), n_extra)
    # NamesMasterList/TeamPOS are Positions' own name-matching contract to the rest of the
    # workbook (NameFix's dropdown, MISC3's NO MATCH check, every VLOOKUP(...,TeamPOS,...))
    # and, like AllProj/AllProjCats, were never being resized to match this sheet's actual row
    # count -- silently stale (and, now, silently missing the whole manual-extra-names block
    # above) the moment the database's roster size changed from whatever run last set them.
    last_pos_row = 1 + n
    set_defined_name(wb, "NamesMasterList", f"Positions!$A$2:$A${last_pos_row}")
    set_defined_name(wb, "TeamPOS", f"Positions!$A$1:$C${last_pos_row}")

    n = refresh_allprojections_names(wb["AllProjections_S"], master["skater_names"])
    log.info("AllProjections_S: %d skater rows", n)

    n = refresh_export_names(wb["Export"], master["skater_names"])
    log.info("Export: %d skater rows", n)

    n = refresh_allprojections_names(wb["AllProjections_G"], master["goalie_names"])
    log.info("AllProjections_G: %d goalie rows", n)

    n = rebuild_settings_sources(wb["Settings"])
    log.info("Settings: rebuilt source list, %d active source(s), compacted to rows 4-%d", n, 3 + n)

    rebuild_misc_source_weights(wb)
    log.info("misc: rebuilt UsedSRCsWTs as a live %d-cell weight lookup (was a dead 12-cell spill)", len(ACTIVE_SOURCES))

    last_col = rebuild_all_projections(wb, "AllProjections_S", SKATER_AP_CATEGORIES, SKATER_AP_CONDITIONAL_TARGETS)
    set_defined_name(wb, "AllProj", f"AllProjections_S!$A:${cl(last_col)}")
    # AllProjCats/AllProjNames are the row-4-header / column-A-name ranges every Player
    # Values MATCH() lookup keys off of -- unlike AllProj above, nothing was resizing these
    # after rebuild_all_projections started varying the sheet's width with ACTIVE_SOURCES,
    # so they stayed pinned at whatever range they last had under the old (wider, 12-source)
    # layout. That silently truncated MATCH's search area mid-category-block: any category
    # landing past the stale cutoff (confirmed live: SHP/SOG/FOW/FOL/HIT/BLK for skaters)
    # could never be found, leaving every Player Values row blank in that column -- while
    # earlier categories (G/A/PIM/PPP, well inside the stale range) kept working, masking the
    # bug as "some columns" rather than "everything broke". Resized here every run so the
    # lookup range always matches the sheet rebuild_all_projections just produced.
    set_defined_name(wb, "AllProjCats", f"AllProjections_S!$A$4:${cl(last_col)}$4")
    set_defined_name(wb, "AllProjNames", f"AllProjections_S!$A$1:$A${wb['AllProjections_S'].max_row}")
    log.info("AllProjections_S: rebuilt, %d categories, ends at column %s", len(SKATER_AP_CATEGORIES), cl(last_col))

    last_col = rebuild_all_projections(wb, "AllProjections_G", GOALIE_AP_CATEGORIES, GOALIE_AP_CONDITIONAL_TARGETS)
    set_defined_name(wb, "AllGProj", f"AllProjections_G!$A:${cl(last_col)}")
    set_defined_name(wb, "AllGProjCats", f"AllProjections_G!$A$4:${cl(last_col)}$4")
    set_defined_name(wb, "AllGProjNames", f"AllProjections_G!$A$1:$A${wb['AllProjections_G'].max_row}")
    log.info("AllProjections_G: rebuilt, %d categories, ends at column %s", len(GOALIE_AP_CATEGORIES), cl(last_col))

    goalie_set = master["goalie_names"]
    first_row, last_row = rebuild_player_values(
        wb["Player Values - Cats"], all_names, goalie_set, "AL")
    log.info("Player Values - Cats: rows %d-%d", first_row, last_row)
    set_defined_name(wb, "RecCats", f"'Player Values - Cats'!$A$2:$H${last_row}")
    set_defined_name(wb, "ValsAll", f"'Player Values - Cats'!$D$2:$G${last_row}")

    first_row, last_row = rebuild_player_values(
        wb["Player Values - Pts"], all_names, goalie_set, "AM")
    log.info("Player Values - Pts: rows %d-%d", first_row, last_row)
    set_defined_name(wb, "RecPoints", f"'Player Values - Pts'!$A$2:$I${last_row}")
    set_defined_name(wb, "FanPtsAll", f"'Player Values - Pts'!$D$2:$G${last_row}")

    n = refresh_rankings(wb["Rankings"], all_names)
    log.info("Rankings: %d rows", n)

    n = rebuild_source_check(wb, all_names)
    log.info("SourceCheck: rebuilt, %d row(s), %d active source column(s)", n, len(ACTIVE_SOURCE_SHEETS))

    n = rebuild_source_comparison(wb["SourceComparison"])
    log.info("SourceComparison: rebuilt, ends at row %d", n)

    rebuild_vorp(wb["CVals/Vorp"], "Player Values - Cats", last_row, "RosterF", "RosterD", "RosterG")
    log.info("CVals/Vorp: rebuilt VorpAll rows 3-%d", last_row)

    rebuild_vorp(wb["FanPts/Vorp"], "Player Values - Pts", last_row, "RosterF", "RosterD", "RosterG")
    log.info("FanPts/Vorp: rebuilt PtsVorpAll rows 3-%d", last_row)

    fh, lh = rebuild_clean_cat(wb["CleanCat"], last_row)
    log.info("CleanCat: hidden block rows %d-%d", fh, lh)
    set_defined_name(wb, "CleanCatF", f"CleanCat!$J$4:$P${3 + 600}")
    set_defined_name(wb, "CleanCatD", f"CleanCat!$R$4:$X${3 + 260}")
    set_defined_name(wb, "CleanCatG", f"CleanCat!$Z$4:$AF${3 + 100}")

    fh, lh = rebuild_clean_pts(wb["CleanPts"], last_row)
    log.info("CleanPts: hidden block rows %d-%d", fh, lh)
    set_defined_name(wb, "CleanPtsF", f"CleanPts!$K$4:$R${3 + 600}")
    set_defined_name(wb, "CleanPtsD", f"CleanPts!$T$4:$AA${3 + 260}")
    set_defined_name(wb, "CleanPtsG", f"CleanPts!$AC$4:$AJ${3 + 100}")

    rebuild_available(wb["Available - Cats"], wb["Player Values - Cats"],
                       "Player Values - Cats", last_row, "G", "H")
    log.info("Available - Cats: rebuilt")

    rebuild_available(wb["Available - Pts"], wb["Player Values - Pts"],
                       "Player Values - Pts", last_row, "G", "I")
    log.info("Available - Pts: rebuilt")

    wb.save()
    log.info("Saved to Google Sheets (spreadsheet id %s)", RESULT_SPREADSHEET_ID)
